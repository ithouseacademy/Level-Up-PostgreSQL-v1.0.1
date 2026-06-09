from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.contrib.auth.hashers import make_password
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.utils import timezone
from django.db.models import Q, Count, Max
from django.db import transaction
from django.core.cache import cache
from django.template.defaulttags import register
import json
import random
from datetime import datetime
from .models import (
    Group, Student, ExamSession, ExamResult, ExamControl,
    AdminPassword, Rules, QuizQuestion, QuizSession, QuizResult,
    Category, GroupCategory, GroupExamConfig, UserExamAttempt,
    CategoryGroupConfig, StudentQuestionHistory,
    ReadingText, ReadingQuestion, StudentAudioPlay,
    Device, Teacher, TeacherScoreLog, AssessmentScore,
    CertificateSetting, Certificate,
    Folder, FolderCategory, GroupFolder, FolderGroupConfig
)
from .forms import GroupForm, RegisterForm, LoginForm


def is_admin_user(user):
    return user.is_authenticated and (user.is_staff or user.is_superuser)


def is_superuser(user):
    return user.is_authenticated and user.is_superuser


@register.filter
def get_item(dictionary, key):
    if dictionary is None:
        return ''
    # q_ prefiksli kalitlarni ham tekshirish
    result = dictionary.get(str(key), '')
    if result == '':
        result = dictionary.get(f'q_{key}', '')
    return result


@register.filter
def get_fill_blank_words(questions):
    if not questions:
        return []
    words = set()
    for q in questions:
        if q.question_type == 'fill_blank':
            for opt in q.get_options_list():
                if opt:
                    words.add(opt.strip())
    return list(words)


def check_answer_correctness(question, user_answer):
    if not user_answer:
        return False
    if isinstance(user_answer, str):
        user_answer = user_answer.strip()
    if not user_answer:
        return False

    user_clean = str(user_answer).strip().lower()

    if question.question_type == 'fill_blank':
        if '|' in question.correct_answer:
            variants = [v.strip().lower() for v in question.correct_answer.split('|') if v.strip()]
            return user_clean in variants
        return user_clean == question.correct_answer.strip().lower()

    elif question.question_type == 'sentence_arrangement':
        user_sentence = ' '.join(user_answer.split()).strip().lower()
        correct_sentence = ' '.join(question.correct_sentence.split()).strip().lower()
        return user_sentence == correct_sentence

    elif question.question_type == 'matching':
        try:
            correct = json.loads(question.correct_answer)
            if isinstance(user_answer, dict):
                for k, v in correct.items():
                    user_val = user_answer.get(k, '')
                    if str(user_val).strip().lower() != str(v).strip().lower():
                        return False
                return True
        except:
            pass
        return False

    else:
        return user_clean == question.correct_answer.strip().lower()


def get_correct_answer_display(question):
    if question.question_type == 'fill_blank':
        if '|' in question.correct_answer:
            return question.correct_answer.split('|')[0].strip()
        return question.correct_answer.strip()
    elif question.question_type == 'sentence_arrangement':
        return question.correct_sentence or ''
    elif question.correct_answer:
        return question.correct_answer
    return ''


def home(request):
    return render(request, 'groups/home.html')


def sayt_haqida(request):
    return render(request, 'groups/sayt_haqida.html')


def user_login(request):
    if request.user.is_authenticated:
        if is_admin_user(request.user):
            return redirect('admin_panel')
        elif is_teacher_user(request.user):
            return redirect('teacher_panel')
        return redirect('student_panel')

    if request.method == 'POST':
        form = LoginForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'Xush kelibsiz, {user.get_full_name() or user.username}!')
                if is_admin_user(user):
                    return redirect('admin_panel')
                elif is_teacher_user(user):
                    return redirect('teacher_panel')
                return redirect('student_panel')
            else:
                messages.error(request, 'Username yoki parol xato!')
        else:
            messages.error(request, 'Username yoki parol xato!')
    else:
        form = LoginForm()

    return render(request, 'groups/login.html', {'form': form, 'title': 'Tizimga kirish'})


def user_register(request):
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    user = User.objects.create_user(
                        username=form.cleaned_data['username'],
                        password=form.cleaned_data['password1'],
                        first_name=form.cleaned_data['first_name'],
                        last_name=form.cleaned_data['last_name']
                    )
                    group = form.cleaned_data['group']
                    student, created = Student.objects.get_or_create(user=user)
                    student.group = group
                    student.save()
                    messages.success(request, 'Tabriklaymiz! Siz muvaffaqiyatli ro\'yxatdan o\'tdingiz!')
                    login(request, user)
                    return redirect('student_panel')
            except Exception as e:
                messages.error(request, f'Xatolik yuz berdi: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{error}')
    else:
        form = RegisterForm()

    return render(request, 'groups/register.html', {'form': form, 'title': 'Ro\'yxatdan o\'tish'})


def user_logout(request):
    logout(request)
    messages.info(request, 'Tizimdan chiqdingiz!')
    return redirect('home')


@login_required
@user_passes_test(is_admin_user)
def admin_panel(request):
    groups = Group.objects.all()
    students = Student.objects.all().select_related('user', 'group')
    admins = User.objects.filter(Q(is_staff=True) | Q(is_superuser=True)).distinct()

    for group in groups:
        if not hasattr(group, 'exam_config'):
            GroupExamConfig.objects.create(group=group)
        if not hasattr(group, 'exam_control'):
            ExamControl.objects.create(group=group)

    context = {
        'groups': groups,
        'students': students,
        'total_groups': groups.count(),
        'total_students': students.count(),
        'total_admins': admins.count(),
        'admins': admins,
        'total_categories': Category.objects.count(),
        'total_questions': QuizQuestion.objects.count(),
    }
    return render(request, 'groups/admin_panel.html', context)


@login_required
def group_detail(request, pk):
    group = get_object_or_404(Group, pk=pk)
    if not is_admin_user(request.user):
        messages.error(request, 'Sizda bu sahifani ko\'rish huquqi yo\'q!')
        return redirect('home')

    students = group.students.all().select_related('user')
    group_categories = GroupCategory.objects.filter(group=group, is_active=True).select_related('category')
    category_ids = group_categories.values_list('category_id', flat=True)
    total_questions = QuizQuestion.objects.filter(category_id__in=category_ids).count()
    quiz_sessions = QuizSession.objects.filter(group=group)
    completed_exams = QuizResult.objects.filter(quiz_session__in=quiz_sessions).count()
    exam_control, created = ExamControl.objects.get_or_create(group=group)
    config, _ = GroupExamConfig.objects.get_or_create(group=group)

    context = {
        'group': group,
        'students': students,
        'group_categories': group_categories,
        'total_questions': total_questions,
        'completed_exams': completed_exams,
        'exam_control': exam_control,
        'config': config,
    }
    return render(request, 'groups/group_detail.html', context)


@login_required
@user_passes_test(is_admin_user)
def group_add(request):
    if request.method == 'POST':
        form = GroupForm(request.POST)
        if form.is_valid():
            group = form.save()
            GroupExamConfig.objects.get_or_create(group=group)
            ExamControl.objects.get_or_create(group=group)
            messages.success(request, f'"{group.name}" guruhi muvaffaqiyatli qo\'shildi!')
            return redirect('admin_panel')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        form = GroupForm()
    return render(request, 'groups/group_form.html', {'form': form, 'title': 'Guruh qo\'shish'})


@login_required
@user_passes_test(is_admin_user)
def group_edit(request, pk):
    group = get_object_or_404(Group, pk=pk)
    if request.method == 'POST':
        form = GroupForm(request.POST, instance=group)
        if form.is_valid():
            form.save()
            messages.success(request, f'"{group.name}" guruhi tahrirlandi!')
            return redirect('admin_panel')
        else:
            for error in form.errors.values():
                messages.error(request, error)
    else:
        form = GroupForm(instance=group)
    return render(request, 'groups/group_form.html', {'form': form, 'title': 'Guruhni tahrirlash'})


@login_required
@user_passes_test(is_admin_user)
def group_delete(request, pk):
    group = get_object_or_404(Group, pk=pk)
    if request.method == 'POST':
        group_name = group.name
        group.delete()
        messages.success(request, f'"{group_name}" guruhi o\'chirildi!')
        return redirect('admin_panel')
    return render(request, 'groups/group_confirm_delete.html', {'group': group})


@login_required
@user_passes_test(is_admin_user)
def student_list(request):
    query = request.GET.get('q', '').strip()
    archive_filter = request.GET.get('archive', '0')
    show_archived = archive_filter == '1'

    if show_archived:
        students = Student.objects.filter(is_archived=True).select_related('user', 'group')
    else:
        students = Student.objects.filter(is_archived=False).select_related('user', 'group')

    if query:
        students = students.filter(
            Q(user__username__icontains=query) |
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(user__email__icontains=query) |
            Q(group__name__icontains=query)
        )
    return render(request, 'groups/student_list.html', {
        'students': students, 'query': query, 'show_archived': show_archived
    })


@login_required
@user_passes_test(is_admin_user)
def student_add(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            try:
                with transaction.atomic():
                    user = User.objects.create_user(
                        username=form.cleaned_data['username'],
                        password=form.cleaned_data['password1'],
                        first_name=form.cleaned_data['first_name'],
                        last_name=form.cleaned_data['last_name']
                    )
                    group = form.cleaned_data['group']
                    Student.objects.create(user=user, group=group)
                    messages.success(request, f'{user.get_full_name()} muvaffaqiyatli qo\'shildi!')
                    return redirect('student_list')
            except Exception as e:
                messages.error(request, f'Xatolik: {str(e)}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{error}')
    else:
        form = RegisterForm()
    return render(request, 'groups/student_form.html', {'form': form, 'title': 'Foydalanuvchi qo\'shish'})


@login_required
@user_passes_test(is_admin_user)
def student_edit(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        try:
            user = student.user
            username = request.POST.get('username', '').strip()
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            group_id = request.POST.get('group')

            if not username:
                messages.error(request, 'Username kiritilishi shart!')
                return redirect('student_edit', pk=pk)

            if User.objects.filter(username=username).exclude(id=user.id).exists():
                messages.error(request, f'"{username}" username allaqachon mavjud!')
                return redirect('student_edit', pk=pk)

            user.username = username
            user.first_name = first_name
            user.last_name = last_name
            user.save()

            if group_id:
                student.group = Group.objects.get(id=group_id)
                student.save()

            messages.success(request, f'{user.get_full_name()} tahrirlandi!')
            return redirect('student_list')
        except Exception as e:
            messages.error(request, f'Xatolik: {str(e)}')

    context = {
        'student': student,
        'groups': Group.objects.all(),
    }
    return render(request, 'groups/student_edit.html', context)


@login_required
@user_passes_test(is_admin_user)
def student_detail(request, pk):
    student = get_object_or_404(Student.objects.select_related('user', 'group'), pk=pk)
    results = QuizResult.objects.filter(student=student).order_by('-submitted_at')
    certificates = Certificate.objects.filter(
        quiz_result__student=student
    ).order_by('-generated_at')

    if request.method == 'POST':
        try:
            user = student.user
            username = request.POST.get('username', '').strip()
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            email = request.POST.get('email', '').strip()
            group_id = request.POST.get('group')

            if not username:
                messages.error(request, 'Username kiritilishi shart!')
                return redirect('student_detail', pk=pk)

            if User.objects.filter(username=username).exclude(id=user.id).exists():
                messages.error(request, f'"{username}" username allaqachon mavjud!')
                return redirect('student_detail', pk=pk)

            user.username = username
            user.first_name = first_name
            user.last_name = last_name
            user.email = email
            user.save()

            if group_id:
                student.group = Group.objects.get(id=group_id)
            else:
                student.group = None
            student.save()

            messages.success(request, f'{user.get_full_name()} tahrirlandi!')
            return redirect('student_detail', pk=pk)
        except Exception as e:
            messages.error(request, f'Xatolik: {str(e)}')

    cert_setting = CertificateSetting.objects.filter(is_active=True).first()
    passing_threshold = cert_setting.threshold_percentage if cert_setting else 70
    context = {
        'student': student,
        'results': results,
        'certificates': certificates,
        'groups': Group.objects.all(),
        'passing_threshold': passing_threshold,
    }
    return render(request, 'groups/student_detail.html', context)


@login_required
@user_passes_test(is_admin_user)
def student_delete(request, pk):
    student = get_object_or_404(Student, pk=pk)
    if request.method == 'POST':
        user = student.user
        full_name = user.get_full_name() or user.username
        student.delete()
        user.delete()
        messages.success(request, f'{full_name} o\'chirildi!')
        return redirect('student_list')
    return render(request, 'groups/student_confirm_delete.html', {'student': student})


@login_required
@csrf_exempt
def accept_rules_api(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Faqat POST'})
    try:
        student = request.user.student_profile
        student.rules_accepted_at = timezone.now()
        student.save()
        return JsonResponse({'success': True, 'accepted_at': str(student.rules_accepted_at)})
    except Student.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Student profili topilmadi'})


@login_required
@csrf_exempt
def change_group_api(request):
    if request.method == 'POST':
        try:
            student = request.user.student_profile
            data = json.loads(request.body)
            group_id = data.get('group_id')
            if group_id:
                group = Group.objects.get(id=group_id)
                student.group = group
            else:
                student.group = None
            student.save()
            return JsonResponse({'success': True, 'group_name': student.group.name if student.group else None})
        except Student.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Student profili topilmadi'})
        except Group.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Guruh topilmadi'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    groups = Group.objects.all().values('id', 'name')
    return JsonResponse({'success': True, 'groups': list(groups)})


@login_required
def student_panel(request):
    try:
        student = request.user.student_profile
        rules = Rules.objects.first()

        exam_active = False
        if student.group:
            exam_control, created = ExamControl.objects.get_or_create(group=student.group)
            exam_active = exam_control.is_active

        context = {
            'student': student,
            'rules': rules,
            'exam_active': exam_active,
        }
        return render(request, 'groups/student_panel.html', context)
    except Student.DoesNotExist:
        messages.error(request, 'Profil topilmadi! Iltimos, admin bilan bog\'laning.')
        return redirect('home')
    except Exception as e:
        messages.error(request, f'Xatolik: {str(e)}')
        return redirect('home')


@login_required
@user_passes_test(is_admin_user)
def student_archive(request, pk):
    student = get_object_or_404(Student, pk=pk)
    student.is_archived = True
    student.save()
    messages.success(request, f'{student.full_name} arxivga olindi!')
    return redirect('student_list')


@login_required
@user_passes_test(is_admin_user)
def student_restore(request, pk):
    student = get_object_or_404(Student, pk=pk)
    student.is_archived = False
    student.save()
    messages.success(request, f'{student.full_name} arxivdan chiqarildi!')
    return redirect('student_list')


@login_required
@user_passes_test(is_admin_user)
def student_bulk_delete(request):
    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        if not student_ids:
            messages.warning(request, "Hech qanday foydalanuvchi tanlanmagan!")
            return redirect('student_list')

        students = Student.objects.filter(pk__in=student_ids)
        student_count = students.count()
        users_to_delete = [student.user for student in students]
        students.delete()
        for user in users_to_delete:
            user.delete()

        messages.success(request, f"{student_count} ta foydalanuvchi muvaffaqiyatli o'chirildi!")
    return redirect('student_list')


@login_required
@user_passes_test(is_admin_user)
def student_bulk_archive(request):
    if request.method == 'POST':
        student_ids = request.POST.getlist('student_ids')
        if not student_ids:
            messages.warning(request, "Hech qanday foydalanuvchi tanlanmagan!")
            return redirect('student_list')

        students = Student.objects.filter(pk__in=student_ids)
        student_count = students.count()
        students.update(is_archived=True)

        messages.success(request, f"{student_count} ta foydalanuvchi arxivga olindi!")
    return redirect('student_list')


@login_required
@user_passes_test(is_superuser)
def make_admin(request):
    if request.method == 'POST':
        user_id = request.POST.get('user_id')
        is_superuser_val = request.POST.get('is_superuser') == 'on'
        try:
            user = User.objects.get(id=user_id)
            user.is_staff = True
            if is_superuser_val:
                user.is_superuser = True
            user.save()
            role = "SUPERUSER" if is_superuser_val else "ADMIN"
            messages.success(request, f'{user.get_full_name()} muvaffaqiyatli {role} qilindi!')
        except User.DoesNotExist:
            messages.error(request, 'Foydalanuvchi topilmadi!')
        except Exception as e:
            messages.error(request, f'Xatolik: {str(e)}')
    return redirect('admin_panel')


@login_required
def admin_list(request):
    admins = User.objects.filter(is_staff=True).order_by('-is_superuser', 'username')
    return render(request, 'groups/admin_list.html', {'admins': admins})


@login_required
@user_passes_test(is_superuser)
def remove_admin(request, user_id):
    if request.method != 'POST':
        messages.error(request, 'Faqat POST so\'rov qabul qilinadi!')
        return redirect('admin_list')

    try:
        user = User.objects.get(id=user_id)
        if user.is_superuser and request.user.id == user.id:
            messages.error(request, "O'zingizni superuserlikdan chiqara olmaysiz!")
        else:
            full_name = user.get_full_name() or user.username
            user.is_staff = False
            user.is_superuser = False
            user.save()
            messages.success(request, f'{full_name} admin huquqidan mahrum qilindi!')
    except User.DoesNotExist:
        messages.error(request, 'Foydalanuvchi topilmadi!')
    except Exception as e:
        messages.error(request, f'Xatolik: {str(e)}')
    return redirect('admin_list')


@login_required
@user_passes_test(is_superuser)
def admin_add(request):
    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')
        is_superuser_val = request.POST.get('is_superuser') == 'on'
        email = request.POST.get('email', '').strip()

        errors = []
        if not first_name: errors.append("Ism kiritilishi shart!")
        if not last_name: errors.append("Familiya kiritilishi shart!")
        if not username: errors.append("Username kiritilishi shart!")
        if not password: errors.append("Parol kiritilishi shart!")
        if len(password) < 4: errors.append("Parol kamida 4 ta belgidan iborat bo'lishi kerak!")
        if password != password_confirm: errors.append("Parollar mos kelmadi!")
        if User.objects.filter(username=username).exists(): errors.append(f"'{username}' username allaqachon mavjud!")

        if errors:
            for error in errors:
                messages.error(request, error)
            return render(request, 'groups/admin_add.html')

        try:
            with transaction.atomic():
                admin_user = User.objects.create(
                    username=username, first_name=first_name, last_name=last_name,
                    email=email, password=make_password(password),
                    is_staff=True, is_superuser=is_superuser_val, is_active=True
                )
                AdminPassword.objects.create(user=admin_user, plain_password=password)
                role = "SUPERUSER" if is_superuser_val else "ADMIN"
                messages.success(request, f'{first_name} {last_name} {role} sifatida qo\'shildi!')
                return redirect('admin_list')
        except Exception as e:
            messages.error(request, f'Xatolik: {str(e)}')

    return render(request, 'groups/admin_add.html')


@login_required
@user_passes_test(is_superuser)
def admin_edit(request, admin_id):
    admin = get_object_or_404(User, id=admin_id)
    try:
        admin_pass = AdminPassword.objects.get(user=admin)
        plain_password = admin_pass.plain_password
    except AdminPassword.DoesNotExist:
        plain_password = ''

    if request.method == 'POST':
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        password_confirm = request.POST.get('password_confirm', '')
        is_superuser_val = request.POST.get('is_superuser') == 'on'

        errors = []
        if not username: errors.append("Username kiritilishi shart!")
        if User.objects.filter(username=username).exclude(id=admin.id).exists():
            errors.append(f"'{username}' username allaqachon mavjud!")
        if password and password != password_confirm: errors.append("Parollar mos kelmadi!")
        if password and len(password) < 4: errors.append("Parol kamida 4 ta belgidan iborat bo'lishi kerak!")

        if errors:
            for error in errors:
                messages.error(request, error)
        else:
            try:
                with transaction.atomic():
                    admin.first_name = first_name
                    admin.last_name = last_name
                    admin.username = username
                    admin.email = email
                    admin.is_superuser = is_superuser_val
                    admin.is_staff = True
                    if password:
                        admin.set_password(password)
                        admin_pass_obj, _ = AdminPassword.objects.get_or_create(user=admin)
                        admin_pass_obj.plain_password = password
                        admin_pass_obj.save()
                    admin.save()
                    messages.success(request, f'{admin.get_full_name()} ma\'lumotlari yangilandi!')
                    return redirect('admin_list')
            except Exception as e:
                messages.error(request, f'Xatolik: {str(e)}')

    return render(request, 'groups/admin_edit.html', {'admin': admin, 'plain_password': plain_password})


@login_required
@user_passes_test(is_superuser)
def admin_delete(request, user_id):
    admin_to_delete = get_object_or_404(User, id=user_id)
    if admin_to_delete.id == request.user.id:
        messages.error(request, "O'zingizni o'chira olmaysiz!")
        return redirect('admin_list')
    if admin_to_delete.is_superuser:
        messages.error(request, "Superuserni o'chira olmaysiz!")
        return redirect('admin_list')
    admin_name = admin_to_delete.get_full_name() or admin_to_delete.username
    try:
        admin_to_delete.delete()
        messages.success(request, f'"{admin_name}" admini muvaffaqiyatli o\'chirildi!')
    except Exception as e:
        messages.error(request, f'Xatolik yuz berdi: {str(e)}')
    return redirect('admin_list')


@login_required
@user_passes_test(is_superuser)
def admin_get_plain_password(request, admin_id):
    try:
        admin = User.objects.get(id=admin_id)
        try:
            admin_pass = AdminPassword.objects.get(user=admin)
            plain_password = admin_pass.plain_password
        except AdminPassword.DoesNotExist:
            plain_password = "Parol saqlanmagan"
        return JsonResponse({
            'success': True,
            'admin_id': admin.id,
            'admin_name': admin.get_full_name() or admin.username,
            'username': admin.username,
            'password': plain_password,
            'is_superuser': admin.is_superuser,
        })
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Admin topilmadi!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Xatolik: {str(e)}'})


@login_required
@user_passes_test(is_superuser)
def admin_update_password(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Faqat POST so\'rov!'})
    try:
        admin_id = request.POST.get('admin_id')
        new_password = request.POST.get('password', '').strip()
        if not new_password:
            return JsonResponse({'success': False, 'message': 'Parol kiritilishi shart!'})
        if len(new_password) < 4:
            return JsonResponse({'success': False, 'message': 'Parol kamida 4 belgi bo\'lishi kerak!'})
        admin = User.objects.get(id=admin_id)
        with transaction.atomic():
            admin.set_password(new_password)
            admin.save()
            admin_pass, _ = AdminPassword.objects.get_or_create(user=admin)
            admin_pass.plain_password = new_password
            admin_pass.save()
        return JsonResponse({'success': True, 'message': 'Parol muvaffaqiyatli yangilandi!'})
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Admin topilmadi!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Xatolik: {str(e)}'})


@login_required
@user_passes_test(is_superuser)
def admin_detail_api(request, admin_id):
    try:
        admin = User.objects.get(id=admin_id)
        return JsonResponse({
            'success': True,
            'admin': {
                'id': admin.id,
                'first_name': admin.first_name,
                'last_name': admin.last_name,
                'username': admin.username,
                'email': admin.email,
                'is_superuser': admin.is_superuser,
            }
        })
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Admin topilmadi!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Xatolik: {str(e)}'})


@login_required
@user_passes_test(is_superuser)
def admin_update(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Faqat POST so\'rov qabul qilinadi!'})
    try:
        admin_id = request.POST.get('admin_id')
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        username = request.POST.get('username', '').strip()
        email = request.POST.get('email', '').strip()
        password = request.POST.get('password', '')
        is_superuser_val = request.POST.get('is_superuser') == 'on'

        if not username:
            return JsonResponse({'success': False, 'message': 'Username kiritilishi shart!'})

        admin = User.objects.get(id=admin_id)
        if admin.id == request.user.id and not is_superuser_val:
            return JsonResponse({'success': False, 'message': 'O\'zingizni superuserlikdan chiqara olmaysiz!'})
        if User.objects.filter(username=username).exclude(id=admin_id).exists():
            return JsonResponse({'success': False, 'message': f'"{username}" username allaqachon mavjud!'})

        with transaction.atomic():
            admin.first_name = first_name
            admin.last_name = last_name
            admin.username = username
            admin.email = email
            if password and len(password) >= 4:
                admin.set_password(password)
                admin_pass, _ = AdminPassword.objects.get_or_create(user=admin)
                admin_pass.plain_password = password
                admin_pass.save()
            if request.user.is_superuser:
                admin.is_superuser = is_superuser_val
                admin.is_staff = True
            admin.save()
        return JsonResponse({'success': True, 'message': 'Admin ma\'lumotlari yangilandi!'})
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Admin topilmadi!'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Xatolik: {str(e)}'})




import json
import random
from collections import defaultdict
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.core.cache import cache
from .models import (
    Group, Student, QuizSession, UserExamAttempt, QuizQuestion, QuizResult,
    GroupExamConfig, GroupCategory, CategoryGroupConfig, ExamControl,
    Category, ReadingText
)


@login_required
def quiz_take(request, group_id):
    try:
        student = Student.objects.get(user=request.user)
    except Student.DoesNotExist:
        return redirect('student_panel')

    group = get_object_or_404(Group, id=group_id)

    if student.group_id != group.id:
        return redirect('student_panel')

    # Exam active check
    is_exam_active = False
    try:
        exam_control = ExamControl.objects.get(group=group)
        is_exam_active = exam_control.is_active
    except ExamControl.DoesNotExist:
        pass

    if not is_exam_active:
        cache_key = f'exam_active_{group_id}'
        is_exam_active = cache.get(cache_key, False)

    if not is_exam_active:
        return render(request, 'groups/quiz_take.html', {
            'group': group,
            'student': student,
            'is_exam_active': False,
            'config': GroupExamConfig.objects.filter(group=group).first(),
        })

    # Quiz session
    quiz_session = QuizSession.objects.filter(group=group, is_active=True).first()
    if not quiz_session:
        quiz_session = QuizSession.objects.create(
            group=group, is_active=True, started_at=timezone.now(), created_by=request.user
        )

    config, _ = GroupExamConfig.objects.get_or_create(group=group)
    max_attempts = config.max_attempts

    # Check completed attempts
    completed_attempts = UserExamAttempt.objects.filter(
        student=student, exam_session=quiz_session, is_completed=True
    ).count()

    if completed_attempts >= max_attempts:
        return render(request, 'groups/quiz_take.html', {
            'group': group, 'student': student, 'is_exam_active': True,
            'has_completed': True, 'already_submitted': True,
            'max_attempts': max_attempts, 'attempts_used': completed_attempts,
            'attempts_left': 0, 'config': config,
        })

    # Get or create attempt
    attempt = UserExamAttempt.objects.filter(
        student=student, exam_session=quiz_session, is_completed=False
    ).first()

    # Load questions — preserve order from stored IDs
    if attempt and attempt.selected_questions:
        q_dict = {q.id: q for q in QuizQuestion.objects.filter(id__in=attempt.selected_questions).select_related('category')}
        questions = [q_dict[qid] for qid in attempt.selected_questions if qid in q_dict]
        user_answers = attempt.user_answers or {}
    else:
        questions_list = []
        selected_category_ids = set()

        # 1. Folder orqali kategoriyalarni olish
        group_folders = GroupFolder.objects.filter(group=group, is_active=True).select_related('folder')
        if group_folders.exists():
            for gf in group_folders:
                folder = gf.folder
                try:
                    folder_config = FolderGroupConfig.objects.get(folder=folder, group=group, is_active=True)
                    cats_to_select = folder_config.categories_to_select
                except FolderGroupConfig.DoesNotExist:
                    cats_to_select = 1

                folder_cats = list(FolderCategory.objects.filter(folder=folder).select_related('category'))
                folder_category_objs = [fc.category for fc in folder_cats]
                if folder_config.randomize_categories:
                    random.shuffle(folder_category_objs)
                take_count = min(cats_to_select, len(folder_category_objs))
                for cat in folder_category_objs[:take_count]:
                    selected_category_ids.add(cat.id)
        else:
            # 2. Agar papka bo'lmasa, eski usul: GroupCategory
            group_categories = GroupCategory.objects.filter(group=group, is_active=True).select_related('category')
            for gc in group_categories:
                selected_category_ids.add(gc.category.id)

        # Har bir tanlangan kategoriyadan savol olish
        for cat_id in selected_category_ids:
            category = Category.objects.get(id=cat_id)
            try:
                cat_config = CategoryGroupConfig.objects.get(category=category, group=group, is_active=True)
                questions_per_category = cat_config.questions_count
            except CategoryGroupConfig.DoesNotExist:
                questions_per_category = 3

            all_cat_questions = list(QuizQuestion.objects.filter(category=category))
            total_available = len(all_cat_questions)

            if total_available > 0:
                take_count = min(questions_per_category, total_available)
                selected = random.sample(all_cat_questions, take_count)
                questions_list.extend(selected)

        # Fallback
        if not questions_list:
            fallback_cat = Category.objects.first()
            if fallback_cat:
                fallback = list(QuizQuestion.objects.filter(category=fallback_cat))
                if fallback:
                    questions_list = random.sample(fallback, min(3, len(fallback)))

        questions = questions_list
        if config.random_order:
            random.shuffle(questions)
        
        user_answers = {}
        attempt = UserExamAttempt.objects.create(
            student=student, exam_session=quiz_session,
            selected_questions=[q.id for q in questions], user_answers={},
            attempt_number=completed_attempts + 1, is_completed=False
        )

    TOTAL_QUESTIONS = len(questions)

    # Reading comprehension data
    for question in questions:
        if question.question_type == 'reading_comprehension' and question.reading_text:
            reading_questions = question.reading_text.reading_questions.all().order_by('order')
            sub_answers = {}
            for sq in reading_questions:
                sub_answers[str(sq.id)] = user_answers.get(f'q_{sq.id}', '')
            
            question.reading_data = {
                'title': question.reading_text.title,
                'content': question.reading_text.content,
                'sub_answers': sub_answers,
                'sub_questions': [
                    {
                        'id': q.id,
                        'text': q.question_text,
                        'correct_answer': q.correct_answer,
                        'user_answer': user_answers.get(f'q_{q.id}', '')
                    }
                    for q in reading_questions
                ]
            }

    # Category grouping va Word Bank
    category_dict = defaultdict(list)
    for question in questions:
        category_dict[question.category].append(question)

    category_list = []
    for category, cat_questions in category_dict.items():
        has_fill_blank = any(q.question_type == 'fill_blank' for q in cat_questions)
        
        cat_fill_blank_words = []
        seen = set()
        
        if has_fill_blank:
            for q in cat_questions:
                if q.question_type == 'fill_blank':
                    options = q.get_options_list()
                    for opt in options:
                        if opt and opt.strip() not in seen:
                            seen.add(opt.strip())
                            cat_fill_blank_words.append(opt.strip())
        
        # Agar so'zlar bo'lmasa, correct_answer dan olish
        if has_fill_blank and not cat_fill_blank_words:
            for q in cat_questions:
                if q.question_type == 'fill_blank' and q.correct_answer:
                    if '|' in q.correct_answer:
                        parts = [p.strip() for p in q.correct_answer.split('|') if p.strip()]
                        for p in parts:
                            if p not in seen:
                                seen.add(p)
                                cat_fill_blank_words.append(p)
                    else:
                        if q.correct_answer.strip() not in seen:
                            seen.add(q.correct_answer.strip())
                            cat_fill_blank_words.append(q.correct_answer.strip())
        
        category_list.append({
            'grouper': category,
            'list': cat_questions,
            'has_fill_blank': has_fill_blank,
            'fill_blank_words': cat_fill_blank_words
        })

    # Categories with speaking/writing must come last
    category_list.sort(key=lambda c: 1 if any(q.question_type in ('speaking', 'writing') for q in c['list']) else 0)

    # Global fill_blank_words
    fill_blank_words = []
    seen_words = set()
    for q in questions:
        if q.question_type == 'fill_blank':
            for opt in q.get_options_list():
                if opt and opt.strip() not in seen_words:
                    seen_words.add(opt.strip())
                    fill_blank_words.append(opt.strip())

    return render(request, 'groups/quiz_take.html', {
        'group': group,
        'student': student,
        'is_exam_active': True,
        'has_completed': False,
        'already_submitted': False,
        'questions': questions,
        'category_list': category_list,
        'user_answers': user_answers,
        'total_questions': TOTAL_QUESTIONS,
        'fill_blank_words': fill_blank_words,
        'attempt_number': attempt.attempt_number,
        'max_attempts': max_attempts,
        'attempts_used': completed_attempts,
        'attempts_left': max_attempts - completed_attempts,
        'config': config,
        'attempt': attempt,
    })


@login_required
@csrf_exempt
def quiz_submit(request):
    """Test javoblarini qabul qilish va baholash"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Faqat POST so\'rov!'})
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'message': 'Avtorizatsiya talab qilinadi'})

    try:
        # JSON ma'lumotlarni o'qish
        try:
            data = json.loads(request.body)
        except:
            data = request.POST.dict()
        
        group_id = data.get('group_id')
        answers = data.get('answers', {})
        
        if isinstance(answers, str):
            try:
                answers = json.loads(answers)
            except:
                answers = {}
        
        student = Student.objects.get(user=request.user)
        group = Group.objects.get(id=group_id)

        # Guruhdagi barcha savollarni olish (folder yoki direct category)
        group_folders = GroupFolder.objects.filter(group=group, is_active=True)
        all_questions = []
        if group_folders.exists():
            for gf in group_folders:
                folder_cats = FolderCategory.objects.filter(folder=gf.folder).values_list('category_id', flat=True)
                for cat_id in folder_cats:
                    questions = list(QuizQuestion.objects.filter(category_id=cat_id))
                    all_questions.extend(questions)
        else:
            group_categories = GroupCategory.objects.filter(group=group, is_active=True)
            for gc in group_categories:
                questions = list(QuizQuestion.objects.filter(category=gc.category))
                all_questions.extend(questions)
        
        config, _ = GroupExamConfig.objects.get_or_create(group=group)
        
        # Sessiyani olish (avval aktiv, topilmasa oxirgi sessiyani)
        quiz_session = QuizSession.objects.filter(group=group, is_active=True).first()
        if not quiz_session:
            quiz_session = QuizSession.objects.filter(group=group).order_by('-started_at').first()
        if not quiz_session:
            quiz_session = QuizSession.objects.create(
                group=group, 
                is_active=False, 
                started_at=timezone.now(), 
                created_by=request.user
            )
        
        # MUHIM: Avval tugallangan attempt bor yoki yo'qligini tekshirish
        existing_completed = UserExamAttempt.objects.filter(
            student=student, 
            exam_session=quiz_session,
            is_completed=True
        ).first()
        
        if existing_completed:
            # Stop_exam_api tomonidan tugallangan bo'lsa, natijani yangilash
            existing_completed.user_answers = answers
            existing_completed.save()
            attempt = existing_completed
        else:
            # Attemptni olish yoki yaratish
            attempt = UserExamAttempt.objects.filter(
                student=student, 
                exam_session=quiz_session,
                is_completed=False
            ).first()
            
            if not attempt:
                attempt = UserExamAttempt.objects.create(
                    student=student,
                    exam_session=quiz_session,
                    selected_questions=[q.id for q in all_questions],
                    user_answers=answers,
                    attempt_number=1,
                    is_completed=False
                )
            else:
                attempt.user_answers = answers
                attempt.save()
        
        # BALLARNI HISOBLASH
        total_score = 0
        total_possible = 0
        question_results = {}
        
        if attempt.selected_questions:
            student_questions = QuizQuestion.objects.filter(id__in=attempt.selected_questions)
        else:
            student_questions = all_questions
        
        for question in student_questions:
            qid = str(question.id)
            
            if question.question_type in ['fill_blank', 'fill_blank_no_word', 'true_false', 'multiple_choice', 'underline_correct']:
                user_answer = answers.get(f'q_{question.id}', '')
                if not user_answer:
                    user_answer = answers.get(str(question.id), '')
                
                is_correct = False
                if user_answer:
                    user_clean = str(user_answer).strip().lower()
                    correct_clean = str(question.correct_answer).strip().lower()
                    
                    if '|' in correct_clean:
                        is_correct = user_clean in [v.strip().lower() for v in correct_clean.split('|')]
                    else:
                        is_correct = user_clean == correct_clean
                
                if is_correct:
                    total_score += question.points
                total_possible += question.points
                
                question_results[qid] = {
                    'user_answer': user_answer if user_answer else 'Javob berilmagan',
                    'is_correct': is_correct,
                    'correct_answer': question.correct_answer
                }
            
            elif question.question_type == 'complete_the_words':
                correct_answers = question.get_complete_words_answers()
                blanks_total = len(correct_answers)
                blanks_correct = 0
                blank_scores = {}
                
                user_answer_raw = answers.get(f'q_{question.id}', '')
                if not user_answer_raw:
                    user_answer_raw = answers.get(str(question.id), '')
                
                user_blank_answers = {}
                if user_answer_raw:
                    if isinstance(user_answer_raw, str):
                        try:
                            user_blank_answers = json.loads(user_answer_raw)
                        except:
                            user_blank_answers = {"1": user_answer_raw}
                    elif isinstance(user_answer_raw, dict):
                        user_blank_answers = user_answer_raw
                
                pts_per_blank = round(question.points / blanks_total, 2)
                for blank_num, correct_val in correct_answers.items():
                    user_val = user_blank_answers.get(str(blank_num), '')
                    if not user_val:
                        user_val = answers.get(f'q_{question.id}_blank_{blank_num}', '')
                    
                    is_correct = False
                    if user_val:
                        user_clean = str(user_val).strip().lower()
                        correct_clean = str(correct_val).strip().lower()
                        if '|' in correct_clean:
                            is_correct = user_clean in [v.strip().lower() for v in correct_clean.split('|')]
                        else:
                            is_correct = user_clean == correct_clean
                    
                    if is_correct:
                        blanks_correct += 1
                        total_score += pts_per_blank
                    
                    blank_scores[str(blank_num)] = {
                        'user_answer': user_val if user_val else 'Javob berilmagan',
                        'is_correct': is_correct,
                        'correct_answer': correct_val
                    }
                
                total_possible += question.points
                question_results[qid] = {
                    'type': 'complete_the_words',
                    'blanks_correct': blanks_correct,
                    'blanks_total': blanks_total,
                    'blanks': blank_scores
                }
            
            elif question.question_type == 'sentence_arrangement':
                user_answer = answers.get(f'q_{question.id}', '')
                correct_sentence = question.correct_sentence or ''
                
                is_correct = False
                if user_answer and correct_sentence:
                    user_clean = ' '.join(str(user_answer).split()).strip().lower()
                    correct_clean = ' '.join(correct_sentence.split()).strip().lower()
                    is_correct = user_clean == correct_clean
                    print(f"SA Check - User: '{user_clean}', Correct: '{correct_clean}', Match: {is_correct}")
                
                if is_correct:
                    total_score += question.points
                total_possible += question.points
                
                question_results[str(question.id)] = {
                    'user_answer': user_answer if user_answer else 'Javob berilmagan',
                    'is_correct': is_correct,
                    'correct_answer': correct_sentence
                }
            elif question.question_type == 'matching':
                correct_answers = question.get_matching_correct_answers()
                blanks_total = len(correct_answers)
                blanks_correct = 0
                blank_scores = {}
                
                pts_per_blank = round(question.points / blanks_total, 2)
                for key, correct_val in correct_answers.items():
                    user_val = answers.get(f'q_{question.id}_{key}', '')
                    if not user_val:
                        user_val = answers.get(f'q_{question.id}', {}).get(str(key), '')
                    
                    is_correct = str(user_val).strip().lower() == str(correct_val).strip().lower()
                    if is_correct:
                        blanks_correct += 1
                        total_score += pts_per_blank
                    
                    blank_scores[key] = {
                        'user_answer': user_val if user_val else 'Javob berilmagan',
                        'is_correct': is_correct,
                        'correct_answer': correct_val
                    }
                
                total_possible += question.points
                question_results[qid] = {
                    'type': 'matching',
                    'blanks_correct': blanks_correct,
                    'blanks_total': blanks_total,
                    'blanks': blank_scores
                }
            
            elif question.question_type == 'cloze_multiple_blanks':
                correct_answers = question.get_cloze_correct_answers()
                blanks_total = len(correct_answers)
                blanks_correct = 0
                blank_scores = {}
                
                user_answers_for_q = answers.get(f'q_{question.id}', {})
                if not isinstance(user_answers_for_q, dict):
                    user_answers_for_q = {}
                
                pts_per_blank = round(question.points / blanks_total, 2)
                for blank_num, correct_val in correct_answers.items():
                    user_val = user_answers_for_q.get(str(blank_num), '')
                    if not user_val:
                        user_val = answers.get(f'q_{question.id}_blank_{blank_num}', '')
                    
                    is_correct = False
                    if user_val:
                        user_clean = str(user_val).strip().lower()
                        correct_clean = str(correct_val).strip().lower()
                        if '|' in correct_clean:
                            is_correct = user_clean in [v.strip().lower() for v in correct_clean.split('|')]
                        else:
                            is_correct = user_clean == correct_clean
                    
                    if is_correct:
                        blanks_correct += 1
                        total_score += pts_per_blank
                    
                    blank_scores[str(blank_num)] = {
                        'user_answer': user_val if user_val else 'Javob berilmagan',
                        'is_correct': is_correct,
                        'correct_answer': correct_val
                    }
                
                total_possible += question.points
                question_results[qid] = {
                    'type': 'cloze_multiple_blanks',
                    'blanks_correct': blanks_correct,
                    'blanks_total': blanks_total,
                    'blanks': blank_scores
                }
            
            elif question.question_type == 'reading_comprehension' and question.reading_text:
                sub_questions = question.reading_text.reading_questions.all()
                blanks_total = sub_questions.count()
                blanks_correct = 0
                blank_scores = {}
                
                pts_per_blank = round(question.points / blanks_total, 2)
                for sq in sub_questions:
                    user_val = answers.get(f'q_{sq.id}', '')
                    is_correct = False
                    if user_val:
                        user_clean = str(user_val).strip().lower()
                        if '|' in sq.correct_answer:
                            is_correct = user_clean in [v.strip().lower() for v in sq.correct_answer.split('|')]
                        else:
                            is_correct = user_clean == sq.correct_answer.strip().lower()
                    
                    if is_correct:
                        blanks_correct += 1
                        total_score += pts_per_blank
                    
                    blank_scores[str(sq.id)] = {
                        'user_answer': user_val if user_val else 'Javob berilmagan',
                        'is_correct': is_correct,
                        'correct_answer': sq.correct_answer
                    }
                
                total_possible += question.points
                question_results[qid] = {
                    'type': 'reading_comprehension',
                    'blanks_correct': blanks_correct,
                    'blanks_total': blanks_total,
                    'blanks': blank_scores
                }

            elif question.question_type == 'writing':
                user_answer = answers.get(f'q_{question.id}', '') or ''
                total_possible += question.points
                question_results[qid] = {
                    'type': 'writing',
                    'user_answer': user_answer,
                    'is_correct': None,
                    'graded': False,
                    'earned_points': 0,
                    'max_points': question.points
                }

            elif question.question_type == 'speaking':
                total_possible += question.points
                question_results[qid] = {
                    'type': 'speaking',
                    'user_answer': '',
                    'is_correct': None,
                    'graded': False,
                    'earned_points': 0,
                    'max_points': question.points
                }

        final_score = 0
        if total_possible > 0:
            final_score = round((total_score / total_possible) * 100, 1)
        
        # Attemptni tugallangan deb belgilash
        attempt.is_completed = True
        attempt.completed_at = timezone.now()
        attempt.user_answers = answers
        attempt.save()
        
        # MUHIM: QuizResult ni faqat bitta marta yaratish
        # get_or_create ishlatamiz, shunda ikki marta yozilmaydi
        result, created = QuizResult.objects.get_or_create(
            student=student, 
            quiz_session=quiz_session,
            attempt_number=attempt.attempt_number,
            defaults={
                'score': final_score,
                'total_questions': total_possible,
                'answers': question_results
            }
        )
        
        if not created:
            # MUHIM: Admin qo'ygan writing/speaking ballarini saqlab qolish
            for qid, existing in result.answers.items():
                if isinstance(existing, dict) and existing.get('type') in ('writing', 'speaking') and existing.get('graded'):
                    if qid in question_results and isinstance(question_results[qid], dict):
                        question_results[qid]['earned_points'] = existing.get('earned_points', 0)
                        question_results[qid]['graded'] = True
                        question_results[qid]['is_correct'] = existing.get('is_correct', existing.get('earned_points', 0) > 0)
                        total_score += existing.get('earned_points', 0)

            final_score = round((total_score / total_possible) * 100, 1) if total_possible > 0 else 0
            result.score = final_score
            result.total_questions = total_possible
            result.answers = question_results
            result.save()
            print(f"Updated existing QuizResult with score: {final_score}")
        else:
            print(f"Created new QuizResult with score: {final_score}")

        try:
            generate_student_certificate(result)
        except Exception as cert_err:
            print(f"Certificate generation error: {cert_err}")

        return JsonResponse({
            'success': True,
            'message': f'Ball: {final_score}/100',
            'score': final_score,
            'total': 100,
            'raw_score': total_score,
            'raw_total': total_possible,
            'already_submitted': False
        })

    except Student.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Student topilmadi!'})
    except Group.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Guruh topilmadi!'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': str(e)})












@login_required
@csrf_exempt
def save_answer_api(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Faqat POST so\'rov'})

    try:
        data = json.loads(request.body)
        student = Student.objects.get(user=request.user)
        group_id = data.get('group_id')

        quiz_session = QuizSession.objects.filter(group_id=group_id, is_active=True).first()
        if not quiz_session:
            # Inactive sessiyada ham saqlashga ruxsat (pauza/stop dan keyin)
            quiz_session = QuizSession.objects.filter(group_id=group_id).order_by('-started_at').first()
            if not quiz_session:
                return JsonResponse({'success': False, 'message': 'Sessiya topilmadi'})

        attempt = UserExamAttempt.objects.filter(
            student=student, exam_session=quiz_session, is_completed=False
        ).first()

        if not attempt:
            # Agar sessiya aktiv bo'lmasa, yangi attempt yaratmaymiz
            if not quiz_session.is_active:
                return JsonResponse({'success': False, 'message': 'Sessiya tugagan'})
            group = Group.objects.get(id=group_id)
            attempt = UserExamAttempt.objects.create(
                student=student,
                exam_session=quiz_session,
                selected_questions=[],
                user_answers={},
                attempt_number=1
            )

        if not attempt.user_answers:
            attempt.user_answers = {}

        # Bulk save (barcha javoblarni bir vaqtda)
        if 'answers' in data:
            for key, value in data['answers'].items():
                if key and value:
                    attempt.user_answers[key] = value
        # Single save (bitta savol javobi)
        elif data.get('question_id'):
            question_id = data.get('question_id')
            answer = data.get('answer', '')
            attempt.user_answers[f'q_{question_id}'] = answer

        attempt.save()

        return JsonResponse({'success': True})

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def get_saved_answers_api(request, group_id):
    try:
        student = Student.objects.get(user=request.user)
        quiz_session = QuizSession.objects.filter(group_id=group_id, is_active=True).first()
        if quiz_session:
            attempt = UserExamAttempt.objects.filter(
                student=student, exam_session=quiz_session, is_completed=False
            ).first()
            if attempt and attempt.user_answers:
                return JsonResponse({'success': True, 'answers': attempt.user_answers})
        return JsonResponse({'success': True, 'answers': {}})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@user_passes_test(is_admin_user)
def quiz_results(request, group_id):
    from django.db.models import Max, Subquery, OuterRef

    group = get_object_or_404(Group, id=group_id)

    # Student o'z natijalarini ko'rishi mumkin
    if not is_admin_user(request.user):
        try:
            student = request.user.student_profile
            if not student.group or student.group.id != group.id:
                messages.error(request, 'Sizda bu sahifani ko\'rish huquqi yo\'q!')
                return redirect('home')
        except:
            messages.error(request, 'Sizda bu sahifani ko\'rish huquqi yo\'q!')
            return redirect('home')

    group_name = group.name

    # Guruh orqali natijalar (guruh o'chirilsa ham saqlangan)
    all_results = QuizResult.objects.none()
    if group:
        all_results = QuizResult.objects.filter(quiz_session__group=group)
    archived = QuizResult.objects.filter(
        quiz_session__isnull=True, group_name_saved=group_name
    )
    all_results = (all_results | archived).distinct().order_by('-submitted_at')

    # Student bo'yicha guruhlash
    unique_results = {}
    for result in all_results:
        student_key = result.student_id or result.student_name_saved
        student_name = result.student_name_saved or (result.student.full_name if result.student else 'Noma\'lum')
        username = result.student.user.username if result.student and result.student.user else ''

        if student_key not in unique_results:
            total_possible = result.total_questions or 0
            score_pct = float(result.score)
            raw_score = round((score_pct / 100) * total_possible, 1) if total_possible > 0 else 0
            unique_results[student_key] = {
                'student_id': result.student_id,
                'student_name': student_name,
                'username': username,
                'last_score': score_pct,
                'raw_score': raw_score,
                'last_submitted': result.submitted_at,
                'total_questions': total_possible,
                'attempt_count': 1,
                'last_result_id': result.id,
            }
        else:
            if result.submitted_at > unique_results[student_key]['last_submitted']:
                total_possible = result.total_questions or 0
                score_pct = float(result.score)
                raw_score = round((score_pct / 100) * total_possible, 1) if total_possible > 0 else 0
                unique_results[student_key]['last_score'] = score_pct
                unique_results[student_key]['raw_score'] = raw_score
                unique_results[student_key]['last_submitted'] = result.submitted_at
                unique_results[student_key]['total_questions'] = total_possible
                unique_results[student_key]['last_result_id'] = result.id
            unique_results[student_key]['attempt_count'] += 1

    # Har bir student uchun sertifikat ID sini olish
    cert_map = {}
    for r in all_results:
        if r.id not in cert_map:
            cert = Certificate.objects.filter(quiz_result=r).first()
            if cert:
                cert_map[r.id] = cert.id
    for item in unique_results.values():
        item['cert_id'] = cert_map.get(item.get('last_result_id'))

    unique_list = list(unique_results.values())
    total_score = sum(r['last_score'] for r in unique_list) if unique_list else 0
    avg_score = round(total_score / len(unique_list), 1) if unique_list else 0
    max_score = max([r['last_score'] for r in unique_list]) if unique_list else 0

    # Har bir student uchun oxirgi qurilmani topish
    for item in unique_list:
        student_id = item['student_id']
        if student_id:
            last_device = Device.objects.filter(student_id=student_id).order_by('-last_seen').first()
            item['device_name'] = last_device.name if last_device and last_device.name else (last_device.device_id[:15] + '...' if last_device else None)
            item['device_platform'] = last_device.platform if last_device else None
        else:
            item['device_name'] = None
            item['device_platform'] = None

    # Baholash tizimi ma'lumoti
    exam_config = GroupExamConfig.objects.filter(group=group).first() if group else None
    for item in unique_list:
        score = item['last_score']
        if exam_config and exam_config.grading_enabled:
            if score < exam_config.low_threshold:
                item['grade'] = exam_config.label_low
                item['grade_class'] = 'red'
            elif score >= exam_config.high_threshold:
                item['grade'] = exam_config.label_high
                item['grade_class'] = 'green'
            else:
                item['grade'] = exam_config.label_medium
                item['grade_class'] = 'yellow'
        else:
            if score >= 70:
                item['grade'] = 'Yuqori'
                item['grade_class'] = 'green'
            elif score >= 50:
                item['grade'] = "O'rta"
                item['grade_class'] = 'yellow'
            else:
                item['grade'] = 'Past'
                item['grade_class'] = 'red'

    context = {
        'group': group,
        'unique_results': unique_list,
        'total_results': all_results.count(),
        'avg_score': avg_score,
        'max_score': max_score,
    }
    return render(request, 'groups/quiz_results.html', context)

@login_required
@user_passes_test(is_admin_user)
@csrf_exempt
@require_http_methods(["POST"])
def start_exam_api(request):
    """Testni boshlash - barcha holatlarni tozalash"""
    try:
        data = json.loads(request.body)
        group_id = data.get('group_id')
        if not group_id:
            return JsonResponse({'success': False, 'message': 'group_id kerak'})

        group = Group.objects.get(id=group_id)
        config, _ = GroupExamConfig.objects.get_or_create(group=group)
        now = timezone.now()

        exam_control, _ = ExamControl.objects.get_or_create(group=group)
        
        # BARCHA HOLATLARNI TOZALASH
        exam_control.is_active = True
        exam_control.is_paused = False
        exam_control.started_at = now
        exam_control.paused_at = None
        exam_control.elapsed_time = 0  # O'tgan vaqtni 0 ga tozalash
        exam_control.save()

        # Cache ni tozalash
        cache.set(f'exam_active_{group_id}', True, timeout=86400)
        cache.delete(f'exam_paused_{group_id}')
        cache.set(f'exam_start_time_{group_id}', now.isoformat(), timeout=86400)
        cache.delete(f'exam_elapsed_time_{group_id}')

        end_time = None
        end_time_str = "Cheksiz"
        if config.time_limit > 0:
            end_time = now + timezone.timedelta(minutes=config.time_limit)
            cache.set(f'exam_end_time_{group_id}', end_time.isoformat(), timeout=86400)
            end_time_str = end_time.strftime('%H:%M:%S')

        # Eski sessiyalarni yopish
        QuizSession.objects.filter(group=group, is_active=True).update(is_active=False)
        ExamSession.objects.filter(group=group, is_active=True).update(is_active=False)

        quiz_session = QuizSession.objects.create(
            group=group, is_active=True, started_at=now, created_by=request.user
        )

        return JsonResponse({
            'success': True,
            'message': f'Test boshlandi! Tugash: {end_time_str}',
            'is_active': True,
            'is_paused': False,
            'elapsed_time': 0,
            'started_at': now.isoformat(),
            'end_time': end_time.isoformat() if end_time else None
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


























def _calculate_score_for_attempt(attempt, group):
    if attempt.selected_questions:
        questions = list(QuizQuestion.objects.filter(id__in=attempt.selected_questions))
    else:
        group_categories = GroupCategory.objects.filter(group=group, is_active=True)
        questions = list(QuizQuestion.objects.filter(
            category__in=group_categories.values('category')
        ))

    user_answers = attempt.user_answers or {}
    total_score = 0
    total_possible = 0
    question_results = {}

    for question in questions:
        qid = str(question.id)

        if question.question_type == 'cloze_multiple_blanks':
            correct_answers = question.get_cloze_correct_answers()
            user_answers_for_q = user_answers.get(f'q_{question.id}', {})
            if not isinstance(user_answers_for_q, dict):
                user_answers_for_q = {}

            blank_scores = {}
            blanks_correct = 0
            blanks_total = len(correct_answers)
            pts_per_blank = round(question.points / blanks_total, 2)

            for blank_num, correct_val in correct_answers.items():
                user_val = user_answers_for_q.get(str(blank_num), '')
                is_blank_correct = False
                if user_val:
                    user_clean = str(user_val).strip().lower()
                    correct_clean = str(correct_val).strip().lower()
                    if '|' in correct_clean:
                        is_blank_correct = user_clean in [v.strip().lower() for v in correct_clean.split('|')]
                    else:
                        is_blank_correct = user_clean == correct_clean

                if is_blank_correct:
                    blanks_correct += 1
                    total_score += pts_per_blank

                blank_scores[str(blank_num)] = {
                    'user_answer': user_val if user_val else 'Javob berilmagan',
                    'is_correct': is_blank_correct,
                    'correct_answer': correct_val
                }

            total_possible += question.points
            question_results[qid] = {
                'type': 'cloze_multiple_blanks',
                'blanks_correct': blanks_correct,
                'blanks_total': blanks_total,
                'blanks': blank_scores
            }

        elif question.question_type == 'matching':
            correct_answers = question.get_matching_correct_answers()
            blanks_total = len(correct_answers)
            blanks_correct = 0
            blank_scores = {}
            pts_per_blank = round(question.points / blanks_total, 2)

            for key, correct_val in correct_answers.items():
                user_val = user_answers.get(f'q_{question.id}_{key}', '')
                is_correct = str(user_val).strip().lower() == str(correct_val).strip().lower()
                if is_correct:
                    blanks_correct += 1
                    total_score += pts_per_blank
                blank_scores[key] = {
                    'user_answer': user_val if user_val else 'Javob berilmagan',
                    'is_correct': is_correct,
                    'correct_answer': correct_val
                }

            total_possible += question.points
            question_results[qid] = {
                'type': 'matching',
                'blanks_correct': blanks_correct,
                'blanks_total': blanks_total,
                'blanks': blank_scores
            }

        elif question.question_type == 'reading_comprehension' and question.reading_text:
            sub_questions = question.reading_text.reading_questions.all()
            blanks_total = sub_questions.count()
            blanks_correct = 0
            blank_scores = {}
            pts_per_blank = round(question.points / blanks_total, 2)

            for sq in sub_questions:
                user_val = user_answers.get(f'q_{sq.id}', '')
                is_correct = False
                if user_val:
                    if '|' in sq.correct_answer:
                        variants = [v.strip().lower() for v in sq.correct_answer.split('|')]
                        is_correct = str(user_val).strip().lower() in variants
                    else:
                        is_correct = str(user_val).strip().lower() == sq.correct_answer.strip().lower()
                if is_correct:
                    blanks_correct += 1
                    total_score += pts_per_blank
                blank_scores[str(sq.id)] = {
                    'user_answer': user_val if user_val else 'Javob berilmagan',
                    'is_correct': is_correct,
                    'correct_answer': sq.correct_answer
                }

            total_possible += question.points
            question_results[qid] = {
                'type': 'reading_comprehension',
                'blanks_correct': blanks_correct,
                'blanks_total': blanks_total,
                'blanks': blank_scores
            }

        elif question.question_type == 'writing':
            total_possible += question.points
            q_key = f'q_{question.id}'
            user_val = user_answers.get(q_key, '') or user_answers.get(str(question.id), '') or ''
            question_results[qid] = {
                'type': 'writing',
                'user_answer': user_val if isinstance(user_val, str) else '',
                'is_correct': None,
                'graded': False,
                'earned_points': 0,
                'max_points': question.points
            }

        elif question.question_type == 'speaking':
            total_possible += question.points
            question_results[qid] = {
                'type': 'speaking',
                'user_answer': '',
                'is_correct': None,
                'graded': False,
                'earned_points': 0,
                'max_points': question.points
            }

        else:
            user_answer = user_answers.get(f'q_{question.id}', '')
            if not user_answer:
                user_answer = user_answers.get(str(question.id), '')

            is_correct = check_answer_correctness(question, user_answer) if user_answer else False
            if is_correct:
                total_score += question.points
            total_possible += question.points

            question_results[qid] = {
                'user_answer': user_answer if user_answer else 'Javob berilmagan',
                'is_correct': is_correct,
                'correct_answer': get_correct_answer_display(question)
            }

    return total_score, total_possible, question_results




































@login_required
@user_passes_test(is_admin_user)
@csrf_exempt
@require_http_methods(["POST"])
def stop_exam_api(request):
    """Testni to'xtatish - faqat tugallanmagan studentlarning javoblarini yig'ish"""
    try:
        data = json.loads(request.body)
        group_id = data.get('group_id')
        if not group_id:
            return JsonResponse({'success': False, 'message': 'group_id kerak'})

        group = Group.objects.get(id=group_id)
        config, _ = GroupExamConfig.objects.get_or_create(group=group)
        
        # BARCHA STUDENTLARNI OLISH
        all_students = Student.objects.filter(group=group)
        
        # Joriy sessiyani olish (avval aktiv, topilmasa oxirgi sessiyani)
        quiz_session = QuizSession.objects.filter(group=group, is_active=True).first()
        if not quiz_session:
            quiz_session = QuizSession.objects.filter(group=group).order_by('-started_at').first()
        if not quiz_session:
            quiz_session = QuizSession.objects.create(
                group=group, 
                is_active=False, 
                started_at=timezone.now(), 
                ended_at=timezone.now(),
                created_by=request.user
            )
        
        # Guruhdagi barcha savollarni olish (ball hisoblash uchun)
        group_folders = GroupFolder.objects.filter(group=group, is_active=True)
        all_questions = []
        if group_folders.exists():
            for gf in group_folders:
                folder_cats = FolderCategory.objects.filter(folder=gf.folder).values_list('category_id', flat=True)
                for cat_id in folder_cats:
                    cat_questions = list(QuizQuestion.objects.filter(category_id=cat_id))
                    all_questions.extend(cat_questions)
        else:
            group_categories = GroupCategory.objects.filter(group=group, is_active=True)
            for gc in group_categories:
                cat_questions = list(QuizQuestion.objects.filter(category=gc.category))
                all_questions.extend(cat_questions)
        
        saved_count = 0
        updated_count = 0
        results_details = []
        
        # HAR BIR STUDENT UCHUN - FAQAT TUGALLANMAGANLARNI QAYTA ISHLASH
        for student in all_students:
            try:
                # MUHIM: Faqat tugallanmagan attemptlarni olish
                attempt = UserExamAttempt.objects.filter(
                    student=student, 
                    exam_session=quiz_session, 
                    is_completed=False
                ).order_by('-attempt_number').first()
                
                # Agar tugallanmagan attempt bo'lmasa, o'tkazib yuboramiz
                if not attempt:
                    # Student allaqachon topshirgan
                    results_details.append({
                        'student': student.full_name,
                        'status': 'already_completed'
                    })
                    continue
                
                # Studentning javoblarini olish
                user_answers = attempt.user_answers or {}
                
                # Studentning savollarini olish
                if attempt.selected_questions:
                    student_questions = QuizQuestion.objects.filter(id__in=attempt.selected_questions)
                else:
                    student_questions = all_questions
                
                # BALLARNI HISOBLASH
                total_score, total_possible, question_results = _calculate_score_for_attempt(attempt, group)
                
                # 100 BALLIK TIZIMGA O'TKAZISH
                final_score = 0
                if total_possible > 0:
                    final_score = round((total_score / total_possible) * 100, 1)
                
                # Attemptni tugallangan deb belgilash
                attempt.is_completed = True
                attempt.completed_at = timezone.now()
                attempt.user_answers = user_answers
                attempt.save()
                
                # MUHIM: QuizResult ni faqat mavjud bo'lmasa yaratish
                result, created = QuizResult.objects.get_or_create(
                    student=student,
                    quiz_session=quiz_session,
                    attempt_number=attempt.attempt_number,
                    defaults={
                        'score': final_score,
                        'total_questions': total_possible,
                        'answers': question_results
                    }
                )
                
                if not created:
                    # MUHIM: Admin qo'ygan writing/speaking ballarini saqlab qolish
                    for qid, existing in result.answers.items():
                        if isinstance(existing, dict) and existing.get('type') in ('writing', 'speaking') and existing.get('graded'):
                            if qid in question_results and isinstance(question_results[qid], dict):
                                question_results[qid]['earned_points'] = existing.get('earned_points', 0)
                                question_results[qid]['graded'] = True
                                question_results[qid]['is_correct'] = existing.get('is_correct', existing.get('earned_points', 0) > 0)
                                total_score += existing.get('earned_points', 0)

                    final_score = round((total_score / total_possible) * 100, 1) if total_possible > 0 else 0
                    result.score = final_score
                    result.total_questions = total_possible
                    result.answers = question_results
                    result.save()
                    updated_count += 1
                else:
                    saved_count += 1
                
                results_details.append({
                    'student': student.full_name,
                    'score': final_score,
                    'total': total_possible,
                    'status': 'completed'
                })
                    
            except Exception as e:
                print(f"Xatolik {student.full_name} uchun: {str(e)}")
                import traceback
                traceback.print_exc()
                results_details.append({
                    'student': student.full_name,
                    'status': f'error: {str(e)}'
                })
        
        # QuizSessionni yopish
        quiz_session.is_active = False
        quiz_session.ended_at = timezone.now()
        quiz_session.save()
        
        # ExamControl ni o'chirish
        ExamControl.objects.filter(group=group).update(
            is_active=False,
            is_paused=False,
            paused_at=None,
            elapsed_time=0
        )
        
        # Cache ni tozalash
        cache.delete(f'exam_active_{group_id}')
        cache.delete(f'exam_start_time_{group_id}')
        cache.delete(f'exam_end_time_{group_id}')
        cache.delete(f'exam_paused_{group_id}')
        cache.delete(f'exam_elapsed_time_{group_id}')
        
        return JsonResponse({
            'success': True,
            'message': f'Test to\'xtatildi! {saved_count} ta yangi natija saqlandi, {updated_count} ta natija yangilandi.',
            'is_active': False,
            'saved_count': saved_count,
            'updated_count': updated_count,
            'total_students': all_students.count(),
            'details': results_details
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': str(e)})











@login_required
@csrf_exempt
@require_http_methods(["POST"])
def auto_stop_exam_api(request):
    try:
        data = json.loads(request.body) if request.body else {}
        group_id = data.get('group_id')
        if not group_id:
            return JsonResponse({'success': False, 'message': 'group_id kerak'})

        group = Group.objects.get(id=group_id)
        quiz_session = QuizSession.objects.filter(group=group, is_active=True).first()

        saved_count = 0

        if quiz_session:
            active_attempts = UserExamAttempt.objects.filter(
                exam_session=quiz_session, is_completed=False
            ).select_related('student')

            for attempt in active_attempts:
                score, total, question_results = _calculate_score_for_attempt(attempt, group)

                attempt.is_completed = True
                attempt.completed_at = timezone.now()
                attempt.save()

                QuizResult.objects.create(
                    student=attempt.student,
                    quiz_session=quiz_session,
                    score=score,
                    total_questions=total,
                    answers=question_results,
                    attempt_number=attempt.attempt_number
                )
                saved_count += 1

            quiz_session.is_active = False
            quiz_session.ended_at = timezone.now()
            quiz_session.save()

        ExamControl.objects.filter(group=group).update(is_active=False)
        cache.delete(f'exam_active_{group_id}')
        cache.delete(f'exam_start_time_{group_id}')
        cache.delete(f'exam_end_time_{group_id}')

        return JsonResponse({
            'success': True,
            'message': f'Vaqt tugadi! {saved_count} ta natija saqlandi.',
            'is_active': False
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def check_time_expired_api(request):
    try:
        data = json.loads(request.body)
        group_id = data.get('group_id')
        if not group_id:
            return JsonResponse({'success': False, 'message': 'group_id kerak'})

        group = Group.objects.get(id=group_id)
        config, _ = GroupExamConfig.objects.get_or_create(group=group)

        if config.time_limit <= 0:
            return JsonResponse({'success': True, 'expired': False})

        start_time_str = cache.get(f'exam_start_time_{group_id}')
        if not start_time_str:
            return JsonResponse({'success': True, 'expired': False})

        start_time = datetime.fromisoformat(start_time_str)
        if timezone.is_naive(start_time):
            start_time = timezone.make_aware(start_time)

        end_time = start_time + timezone.timedelta(minutes=config.time_limit)

        if timezone.now() >= end_time:
            return auto_stop_exam_api(request)

        return JsonResponse({'success': True, 'expired': False})

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def exam_control(request, group_id):
    if not is_admin_user(request.user):
        messages.error(request, 'Sizda bu sahifani ko\'rish huquqi yo\'q!')
        return redirect('home')

    group = get_object_or_404(Group, id=group_id)
    exam_control_obj, created = ExamControl.objects.get_or_create(group=group)
    students = group.students.all().select_related('user')
    config, _ = GroupExamConfig.objects.get_or_create(group=group)

    group_categories = GroupCategory.objects.filter(group=group).values_list('category_id', flat=True)
    questions_count = QuizQuestion.objects.filter(category_id__in=group_categories).count()

    context = {
        'group': group,
        'exam_control': exam_control_obj,
        'students': students,
        'questions_count': questions_count,
        'config': config,
    }
    return render(request, 'groups/exam_control.html', context)


@login_required
@csrf_exempt
def check_audio_play_api(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Faqat POST so\'rov'})

    try:
        data = json.loads(request.body)
        group_id = data.get('group_id')
        category_id = data.get('category_id')

        if not group_id or not category_id:
            return JsonResponse({'success': False, 'message': 'group_id va category_id kerak'})

        student = Student.objects.get(user=request.user)
        group = Group.objects.get(id=group_id)
        category = Category.objects.get(id=category_id)

        quiz_session = QuizSession.objects.filter(group=group, is_active=True).first()

        max_plays = category.max_audio_plays if category.max_audio_plays is not None else 1

        audio_play, created = StudentAudioPlay.objects.get_or_create(
            student=student,
            group=group,
            category=category,
            exam_session=quiz_session,
            defaults={'max_plays': max_plays, 'play_count': 0}
        )

        if audio_play.max_plays != max_plays:
            audio_play.max_plays = max_plays
            audio_play.save()

        return JsonResponse({
            'success': True,
            'can_play': audio_play.can_play(),
            'play_count': audio_play.play_count,
            'max_plays': audio_play.max_plays
        })

    except Student.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Student topilmadi!'})
    except Group.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Guruh topilmadi!'})
    except Category.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Kategoriya topilmadi!'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@csrf_exempt
def record_audio_play_api(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Faqat POST so\'rov'})

    try:
        data = json.loads(request.body)
        group_id = data.get('group_id')
        category_id = data.get('category_id')

        if not group_id or not category_id:
            return JsonResponse({'success': False, 'message': 'group_id va category_id kerak'})

        student = Student.objects.get(user=request.user)
        group = Group.objects.get(id=group_id)
        category = Category.objects.get(id=category_id)

        quiz_session = QuizSession.objects.filter(group=group, is_active=True).first()

        max_plays = category.max_audio_plays if category.max_audio_plays is not None else 1

        audio_play, created = StudentAudioPlay.objects.get_or_create(
            student=student,
            group=group,
            category=category,
            exam_session=quiz_session,
            defaults={'max_plays': max_plays, 'play_count': 0}
        )

        if audio_play.can_play():
            audio_play.increment_play()
            max_display = audio_play.max_plays if audio_play.max_plays > 0 else '∞'
            return JsonResponse({
                'success': True,
                'play_count': audio_play.play_count,
                'can_play': audio_play.can_play(),
                'max_plays': audio_play.max_plays,
                'message': f'Audio eshitildi ({audio_play.play_count}/{max_display})'
            })
        else:
            return JsonResponse({
                'success': False,
                'message': f'Siz maksimal {audio_play.max_plays} marta eshitgansiz!'
            })

    except Student.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Student topilmadi!'})
    except Group.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Guruh topilmadi!'})
    except Category.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Kategoriya topilmadi!'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@user_passes_test(is_admin_user)
def quiz_admin(request):
    questions = QuizQuestion.objects.all().select_related('category').order_by('-created_at')
    categories = Category.objects.all()
    groups = Group.objects.all()
    return render(request, 'groups/quiz_admin.html', {
        'questions': questions,
        'categories': categories,
        'groups': groups,
    })


@login_required
@user_passes_test(is_admin_user)
def quiz_add_question(request):
    if request.method == 'POST':
        category_id = request.POST.get('category_id')
        question_text = request.POST.get('question_text', '').strip()
        correct_answer = request.POST.get('correct_answer', '').strip().lower()

        if not question_text or not correct_answer:
            messages.error(request, 'Savol matni va to\'g\'ri javob kiritilishi shart!')
            return redirect('quiz_admin')

        if category_id:
            try:
                category = Category.objects.get(id=category_id)
                QuizQuestion.objects.create(
                    category=category,
                    question_text=question_text,
                    correct_answer=correct_answer
                )
                messages.success(request, f'Savol "{category.name}" kategoriyasiga qo\'shildi!')
            except Category.DoesNotExist:
                messages.error(request, 'Kategoriya topilmadi!')
        else:
            messages.error(request, 'Kategoriya tanlash shart!')
    return redirect('quiz_admin')


@login_required
@user_passes_test(is_admin_user)
def quiz_edit_question(request, question_id):
    question = get_object_or_404(QuizQuestion, id=question_id)
    if request.method == 'POST':
        category_id = request.POST.get('category_id')
        question_text = request.POST.get('question_text', '').strip()
        correct_answer = request.POST.get('correct_answer', '').strip().lower()
        if not question_text or not correct_answer:
            messages.error(request, 'Savol matni va to\'g\'ri javob kiritilishi shart!')
        else:
            try:
                question.question_text = question_text
                question.correct_answer = correct_answer
                if category_id:
                    question.category = Category.objects.get(id=category_id)
                question.save()
                messages.success(request, 'Savol muvaffaqiyatli tahrirlandi!')
            except Category.DoesNotExist:
                messages.error(request, 'Kategoriya topilmadi!')
            except Exception as e:
                messages.error(request, f'Xatolik: {str(e)}')
        return redirect('quiz_admin')

    return render(request, 'groups/quiz_edit.html', {
        'question': question,
        'categories': Category.objects.all(),
    })


@login_required
@user_passes_test(is_admin_user)
def quiz_delete_question(request, question_id):
    question = get_object_or_404(QuizQuestion, id=question_id)
    question.delete()
    messages.success(request, 'Savol o\'chirildi!')
    return redirect('quiz_admin')

@login_required
@user_passes_test(is_admin_user)
def quiz_result_details_api(request, result_id):
    try:
        result = QuizResult.objects.get(id=result_id)
        answers_html = '<div class="space-y-2">'
        
        if result.answers:
            for qid, answer in result.answers.items():
                if isinstance(answer, dict):
                    # Agar 'blanks' mavjud bo'lsa (matching, cloze, reading)
                    if 'blanks' in answer:
                        for bnum, bdata in answer['blanks'].items():
                            is_correct = bdata.get('is_correct', False)
                            answers_html += f'''
                            <div class="border-l-4 {'border-green-500' if is_correct else 'border-red-400'} bg-gray-50 p-3 rounded">
                                <p class="text-sm">
                                    <span class="text-gray-500">Savol {qid} - Bo\'sh joy {bnum}:</span><br>
                                    <span class="font-medium">Javob: {bdata.get('user_answer', '-')}</span>
                                    {'✅' if is_correct else '❌'}
                                    <span class="text-gray-400 text-xs ml-2">(To\'g\'ri: {bdata.get('correct_answer', '-')})</span>
                                </p>
                            </div>'''
                    else:
                        # Oddiy savol
                        is_correct = answer.get('is_correct', False)
                        user_answer = answer.get('user_answer', 'Javob berilmagan')
                        correct_answer = answer.get('correct_answer', '-')
                        answers_html += f'''
                        <div class="border-l-4 {'border-green-500' if is_correct else 'border-red-400'} bg-gray-50 p-3 rounded">
                            <p class="text-sm">
                                <span class="text-gray-500">Savol {qid}:</span><br>
                                <span class="font-medium">Javob: {user_answer}</span>
                                {'✅' if is_correct else '❌'}
                                <span class="text-gray-400 text-xs ml-2">(To\'g\'ri: {correct_answer})</span>
                            </p>
                        </div>'''
                else:
                    # String formatdagi javob
                    answers_html += f'''
                    <div class="border-l-4 border-gray-400 bg-gray-50 p-3 rounded">
                        <p class="text-sm">Savol {qid}: {answer}</p>
                    </div>'''
        else:
            answers_html += '<p class="text-gray-500 text-center">Javoblar tafsilotlari mavjud emas</p>'
        
        answers_html += '</div>'

        total_possible = result.total_questions or 0
        score_pct = float(result.score)
        raw_score = round((score_pct / 100) * total_possible, 1) if total_possible > 0 else 0
        return JsonResponse({
            'success': True,
            'student_name': result.student.full_name,
            'score': score_pct,
            'raw_score': raw_score,
            'total': total_possible,
            'percentage': float(result.percentage),
            'submitted_at': result.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
            'answers_html': answers_html
        })
    except QuizResult.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Natija topilmadi'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

@login_required
@user_passes_test(is_admin_user)
def group_exam_config(request, group_id):
    group = get_object_or_404(Group, id=group_id)
    config, created = GroupExamConfig.objects.get_or_create(group=group)
    group_categories = GroupCategory.objects.filter(group=group, is_active=True).select_related('category')
    category_ids = [gc.category.id for gc in group_categories]
    cat_ids_from_folders = set()
    group_folders = GroupFolder.objects.filter(group=group, is_active=True).select_related('folder')
    folder_configs = {}
    for gf in group_folders:
        folder = gf.folder
        fconfig, _ = FolderGroupConfig.objects.get_or_create(
            folder=folder, group=group,
            defaults={'categories_to_select': 1, 'randomize_categories': True, 'is_active': True}
        )
        folder_configs[str(folder.id)] = fconfig
        folder_cats = FolderCategory.objects.filter(folder=folder).values_list('category_id', flat=True)
        for cid in folder_cats:
            cat_ids_from_folders.add(cid)
    all_cat_ids = list(set(category_ids) | cat_ids_from_folders)
    total_questions = QuizQuestion.objects.filter(category_id__in=all_cat_ids).count()

    category_configs = {}
    for gc in group_categories:
        cat_config, _ = CategoryGroupConfig.objects.get_or_create(
            category=gc.category, group=group,
            defaults={'questions_count': 3, 'random_order': True, 'is_active': True}
        )
        category_configs[gc.category.id] = cat_config

    if request.method == 'POST':
        try:
            questions_per_student = int(request.POST.get('questions_per_student', 5))
            random_order = request.POST.get('random_order') == 'on'
            show_correct_answer = request.POST.get('show_correct_answer') == 'on'
            time_limit = int(request.POST.get('time_limit', 0))
            max_attempts = int(request.POST.get('max_attempts', 1))
            use_category_configs = request.POST.get('use_category_configs') == 'on'

            # Baholash tizimi sozlamalari
            grading_enabled = request.POST.get('grading_enabled') == 'on'
            low_threshold = int(request.POST.get('low_threshold', 40))
            high_threshold = int(request.POST.get('high_threshold', 70))
            label_low = request.POST.get('label_low', 'Past').strip()
            label_medium = request.POST.get('label_medium', "O'rta").strip()
            label_high = request.POST.get('label_high', 'Yuqori').strip()

            if questions_per_student <= 0:
                messages.error(request, 'Savollar soni 0 dan katta bo\'lishi kerak!')
                return redirect('group_exam_config', group_id=group.id)
            if max_attempts <= 0:
                messages.error(request, 'Urinishlar soni 0 dan katta bo\'lishi kerak!')
                return redirect('group_exam_config', group_id=group.id)

            config.questions_per_student = questions_per_student
            config.random_order = random_order
            config.show_correct_answer = show_correct_answer
            config.time_limit = time_limit
            config.max_attempts = max_attempts
            config.use_category_configs = use_category_configs
            config.grading_enabled = grading_enabled
            config.low_threshold = low_threshold
            config.high_threshold = high_threshold
            config.label_low = label_low
            config.label_medium = label_medium
            config.label_high = label_high

            # Sertifikat sozlamalari
            config.certificate_enabled = request.POST.get('certificate_enabled') == 'on'
            config.certificate_level = request.POST.get('certificate_level', '').strip()
            config.certificate_teacher = request.POST.get('certificate_teacher', '').strip()
            config.save()

            messages.success(request, '✅ Sozlamalar saqlandi!')
            return redirect('group_exam_config', group_id=group.id)
        except ValueError:
            messages.error(request, 'Noto\'g\'ri raqam kiritildi!')
        except Exception as e:
            messages.error(request, f'Xatolik: {str(e)}')

    LEVEL_CHOICES = [ 'English Proficiency Level: A1 (Starter)', 'English Proficiency Level: A1 (Beginner)', 'English Proficiency Level: A2 (Elementary)', 'English Proficiency Level: B1 (Pre-Intermediate)', 'English Proficiency Level: B2 (Intermediate)', 'English Proficiency Level: C1 (Upper-Intermediate)', 'English Proficiency Level: C2 (Advanced)']

    context = {
        'group': group,
        'config': config,
        'total_questions': total_questions,
        'group_categories': group_categories,
        'category_configs': category_configs,
        'level_choices': LEVEL_CHOICES,
        'group_folders': group_folders,
        'folder_configs': folder_configs,
    }
    return render(request, 'groups/group_exam_config.html', context)


@login_required
@user_passes_test(is_admin_user)
@csrf_exempt
def save_category_configs_api(request, group_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Faqat POST so\'rov'})
    try:
        group = get_object_or_404(Group, id=group_id)
        data = json.loads(request.body)
        configs = data.get('configs', [])
        for cfg in configs:
            category_id = cfg.get('category_id')
            questions_count = cfg.get('questions_count', 5)
            random_order = cfg.get('random_order', True)
            CategoryGroupConfig.objects.update_or_create(
                group=group, category_id=category_id,
                defaults={'questions_count': questions_count, 'random_order': random_order, 'is_active': True}
            )
        return JsonResponse({'success': True, 'message': f'{len(configs)} ta sozlama saqlandi'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@user_passes_test(is_admin_user)
def group_questions_preview(request, group_id):
    group = get_object_or_404(Group, id=group_id)
    config, _ = GroupExamConfig.objects.get_or_create(group=group)
    group_folders = GroupFolder.objects.filter(group=group, is_active=True)
    if group_folders.exists():
        cat_ids = set()
        for gf in group_folders:
            folder_cats = FolderCategory.objects.filter(folder=gf.folder).values_list('category_id', flat=True)
            for cid in folder_cats:
                cat_ids.add(cid)
        all_questions = list(QuizQuestion.objects.filter(category_id__in=list(cat_ids)))
    else:
        group_categories = GroupCategory.objects.filter(group=group).values_list('category_id', flat=True)
        all_questions = list(QuizQuestion.objects.filter(category_id__in=group_categories))
    question_count = min(config.questions_per_student, len(all_questions))
    random_questions = random.sample(all_questions, question_count) if question_count > 0 else []

    return render(request, 'groups/group_questions_preview.html', {
        'group': group,
        'config': config,
        'random_questions': random_questions,
        'total_available': len(all_questions),
    })


@login_required
@user_passes_test(is_admin_user)
def category_list(request):
    categories = Category.objects.all().annotate(
        questions_count=Count('quiz_questions'),
        groups_count=Count('group_categories')
    )
    return render(request, 'groups/category_list.html', {
        'categories': categories,
        'total_categories': categories.count(),
    })


@login_required
@user_passes_test(is_admin_user)
def category_add(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        if not name:
            messages.error(request, 'Kategoriya nomi kiritilishi shart!')
        elif Category.objects.filter(name__iexact=name).exists():
            messages.error(request, f'"{name}" nomli kategoriya allaqachon mavjud!')
        else:
            Category.objects.create(name=name, description=description)
            messages.success(request, f'✅ "{name}" kategoriyasi qo\'shildi!')
            return redirect('category_list')
    return render(request, 'groups/category_form.html', {'title': 'Kategoriya qo\'shish'})


@login_required
@user_passes_test(is_admin_user)
def category_edit(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        if not name:
            messages.error(request, 'Kategoriya nomi kiritilishi shart!')
        elif Category.objects.filter(name__iexact=name).exclude(pk=pk).exists():
            messages.error(request, f'"{name}" nomli kategoriya allaqachon mavjud!')
        else:
            category.name = name
            category.description = description
            category.save()
            messages.success(request, f'✅ "{name}" kategoriyasi yangilandi!')
            return redirect('category_list')
    return render(request, 'groups/category_form.html', {
        'category': category,
        'title': 'Kategoriyani tahrirlash'
    })


@login_required
@user_passes_test(is_admin_user)
def category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        questions_count = category.quiz_questions.count()
        category_name = category.name
        category.delete()
        if questions_count > 0:
            messages.warning(request, f'🗑️ "{category_name}" va {questions_count} ta savol o\'chirildi!')
        else:
            messages.success(request, f'🗑️ "{category_name}" kategoriyasi o\'chirildi!')
        return redirect('category_list')
    return render(request, 'groups/category_confirm_delete.html', {'category': category})


@login_required
@user_passes_test(is_admin_user)
def category_group_config(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    configs = CategoryGroupConfig.objects.filter(category=category).select_related('group')
    all_groups = Group.objects.all()
    configured_group_ids = configs.values_list('group_id', flat=True)
    available_groups = all_groups.exclude(id__in=configured_group_ids)

    return render(request, 'groups/category_group_config.html', {
        'category': category,
        'configs': configs,
        'available_groups': available_groups,
        'total_questions': category.quiz_questions.count(),
    })


@login_required
@user_passes_test(is_admin_user)
def category_group_config_add(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    if request.method == 'POST':
        group_id = request.POST.get('group_id')
        questions_count = int(request.POST.get('questions_count', 5))
        random_order = request.POST.get('random_order') == 'on'
        total_questions = category.quiz_questions.count()

        if questions_count > total_questions and total_questions > 0:
            messages.warning(request, f"Kategoriyada faqat {total_questions} ta savol bor.")
            questions_count = total_questions
        if questions_count <= 0:
            messages.error(request, "Savollar soni 0 dan katta bo'lishi kerak!")
            return redirect('category_group_config', category_id=category.id)

        try:
            group = Group.objects.get(id=group_id)
            config, created = CategoryGroupConfig.objects.get_or_create(
                category=category, group=group,
                defaults={'questions_count': questions_count, 'random_order': random_order, 'is_active': True}
            )
            if not created:
                config.questions_count = questions_count
                config.random_order = random_order
                config.save()
                messages.success(request, 'Sozlamalar yangilandi!')
            else:
                messages.success(request, f'{group.name} guruhi uchun sozlamalar qo\'shildi!')
        except Group.DoesNotExist:
            messages.error(request, 'Guruh topilmadi!')
        except Exception as e:
            messages.error(request, f'Xatolik: {str(e)}')
    return redirect('category_group_config', category_id=category.id)


@login_required
@user_passes_test(is_admin_user)
def category_group_config_edit(request, config_id):
    config = get_object_or_404(CategoryGroupConfig, id=config_id)
    if request.method == 'POST':
        questions_count = int(request.POST.get('questions_count', 5))
        random_order = request.POST.get('random_order') == 'on'
        is_active = request.POST.get('is_active') == 'on'
        total_questions = config.category.quiz_questions.count()

        if questions_count > total_questions and total_questions > 0:
            questions_count = total_questions
        if questions_count <= 0:
            messages.error(request, "Savollar soni 0 dan katta bo'lishi kerak!")
        else:
            config.questions_count = questions_count
            config.random_order = random_order
            config.is_active = is_active
            config.save()
            messages.success(request, 'Sozlamalar saqlandi!')
        return redirect('category_group_config', category_id=config.category.id)

    return render(request, 'groups/category_group_config_edit.html', {
        'config': config,
        'total_questions': config.category.quiz_questions.count(),
    })


@login_required
@user_passes_test(is_admin_user)
def category_group_config_delete(request, config_id):
    config = get_object_or_404(CategoryGroupConfig, id=config_id)
    category_id = config.category.id
    group_name = config.group.name
    if request.method == 'POST':
        config.delete()
        messages.success(request, f'{group_name} guruhi uchun sozlamalar o\'chirildi!')
        return redirect('category_group_config', category_id=category_id)
    return render(request, 'groups/category_group_config_confirm_delete.html', {'config': config})


@login_required
@user_passes_test(is_admin_user)
def category_questions_list(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    questions = QuizQuestion.objects.filter(category=category).order_by('id')
    groups_using = Group.objects.filter(group_categories__category=category).distinct()
    return render(request, 'groups/category_questions_list.html', {
        'category': category,
        'questions': questions,
        'total_questions': questions.count(),
        'groups_using': groups_using,
    })


@login_required
@user_passes_test(is_admin_user)
def category_question_add(request, category_id):
    category = get_object_or_404(Category, id=category_id)
    if request.method == 'POST':
        question_text = request.POST.get('question_text', '').strip()
        correct_answer = request.POST.get('correct_answer', '').strip().lower()
        if not question_text or not correct_answer:
            messages.error(request, 'Savol matni va to\'g\'ri javob kiritilishi shart!')
        else:
            QuizQuestion.objects.create(
                category=category, question_text=question_text, correct_answer=correct_answer
            )
            messages.success(request, f'✅ "{category.name}" kategoriyasiga yangi savol qo\'shildi!')
            return redirect('category_questions_list', category_id=category.id)
    return render(request, 'groups/category_question_form.html', {
        'category': category, 'is_edit': False,
    })


@login_required
@user_passes_test(is_admin_user)
def category_question_edit(request, question_id):
    question = get_object_or_404(QuizQuestion, id=question_id)
    category = question.category
    if request.method == 'POST':
        question_text = request.POST.get('question_text', '').strip()
        correct_answer = request.POST.get('correct_answer', '').strip().lower()
        if not question_text or not correct_answer:
            messages.error(request, 'Savol matni va to\'g\'ri javob kiritilishi shart!')
        else:
            question.question_text = question_text
            question.correct_answer = correct_answer
            question.save()
            messages.success(request, '✅ Savol tahrirlandi!')
            return redirect('category_questions_list', category_id=category.id)
    return render(request, 'groups/category_question_form.html', {
        'question': question, 'category': category, 'is_edit': True,
    })


@login_required
@user_passes_test(is_admin_user)
def category_question_delete(request, question_id):
    question = get_object_or_404(QuizQuestion, id=question_id)
    category = question.category
    if request.method == 'POST':
        question.delete()
        messages.success(request, '🗑️ Savol o\'chirildi!')
        return redirect('category_questions_list', category_id=category.id)
    return render(request, 'groups/category_question_confirm_delete.html', {
        'question': question, 'category': category,
    })


@login_required
@user_passes_test(is_admin_user)
def group_categories_manage(request, group_id):
    group = get_object_or_404(Group, id=group_id)
    assigned_categories = GroupCategory.objects.filter(group=group).select_related('category')
    assigned_ids = [gc.category.id for gc in assigned_categories]
    available_categories = Category.objects.exclude(id__in=assigned_ids)
    return render(request, 'groups/group_categories_manage.html', {
        'group': group,
        'assigned_categories': assigned_categories,
        'available_categories': available_categories,
    })


@login_required
@user_passes_test(is_admin_user)
def group_category_add(request, group_id):
    if request.method != 'POST':
        messages.error(request, 'Faqat POST so\'rov!')
        return redirect('group_categories_manage', group_id=group_id)
    try:
        group = get_object_or_404(Group, id=group_id)
        category_id = request.POST.get('category_id')
        if not category_id:
            messages.error(request, 'Kategoriya tanlanmagan!')
            return redirect('group_categories_manage', group_id=group_id)
        category = get_object_or_404(Category, id=category_id)
        if not GroupCategory.objects.filter(group=group, category=category).exists():
            GroupCategory.objects.create(group=group, category=category)
            messages.success(request, f'✅ "{category.name}" kategoriyasi qo\'shildi!')
        else:
            messages.warning(request, f'"{category.name}" allaqachon mavjud!')
    except Exception as e:
        messages.error(request, f'Xatolik: {str(e)}')
    return redirect('group_categories_manage', group_id=group_id)


@login_required
@user_passes_test(is_admin_user)
def group_category_remove(request, group_category_id):
    if request.method != 'POST':
        messages.error(request, 'Faqat POST so\'rov!')
        return redirect('admin_panel')
    try:
        group_category_obj = get_object_or_404(GroupCategory, id=group_category_id)
        group = group_category_obj.group
        category_name = group_category_obj.category.name
        group_category_obj.delete()
        messages.success(request, f'🗑️ "{category_name}" kategoriyasi olib tashlandi!')
    except Exception as e:
        messages.error(request, f'Xatolik: {str(e)}')
        return redirect('admin_panel')
    return redirect('group_categories_manage', group_id=group.id)


# ==================== FOLDER (PAPKA) VIEWS ====================

@login_required
@user_passes_test(is_admin_user)
def folder_list(request):
    folders = Folder.objects.all().annotate(
        categories_count=Count('folder_categories'),
    )
    return render(request, 'groups/folder_list.html', {
        'folders': folders,
        'total_folders': folders.count(),
    })


@login_required
@user_passes_test(is_admin_user)
def folder_add(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        if not name:
            messages.error(request, 'Papka nomi kiritilishi shart!')
        elif Folder.objects.filter(name__iexact=name).exists():
            messages.error(request, f'"{name}" nomli papka allaqachon mavjud!')
        else:
            Folder.objects.create(name=name, description=description)
            messages.success(request, f'✅ "{name}" papkasi qo\'shildi!')
            return redirect('folder_list')
    return render(request, 'groups/folder_form.html', {'title': 'Papka qo\'shish'})


@login_required
@user_passes_test(is_admin_user)
def folder_edit(request, pk):
    folder = get_object_or_404(Folder, pk=pk)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        if not name:
            messages.error(request, 'Papka nomi kiritilishi shart!')
        elif Folder.objects.filter(name__iexact=name).exclude(pk=pk).exists():
            messages.error(request, f'"{name}" nomli papka allaqachon mavjud!')
        else:
            folder.name = name
            folder.description = description
            folder.save()
            messages.success(request, f'✅ "{name}" papkasi yangilandi!')
            return redirect('folder_list')
    return render(request, 'groups/folder_form.html', {
        'folder': folder,
        'title': 'Papkani tahrirlash'
    })


@login_required
@user_passes_test(is_admin_user)
def folder_delete(request, pk):
    folder = get_object_or_404(Folder, pk=pk)
    if request.method == 'POST':
        folder_name = folder.name
        folder.delete()
        messages.success(request, f'🗑️ "{folder_name}" papkasi o\'chirildi!')
        return redirect('folder_list')
    return render(request, 'groups/folder_confirm_delete.html', {'folder': folder})


@login_required
@user_passes_test(is_admin_user)
def folder_categories_manage(request, folder_id):
    folder = get_object_or_404(Folder, id=folder_id)
    assigned_categories = FolderCategory.objects.filter(folder=folder).select_related('category')
    assigned_ids = [fc.category.id for fc in assigned_categories]
    available_categories = Category.objects.exclude(id__in=assigned_ids)
    return render(request, 'groups/folder_categories_manage.html', {
        'folder': folder,
        'assigned_categories': assigned_categories,
        'available_categories': available_categories,
    })


@login_required
@user_passes_test(is_admin_user)
def folder_category_add(request, folder_id):
    if request.method != 'POST':
        messages.error(request, 'Faqat POST so\'rov!')
        return redirect('folder_categories_manage', folder_id=folder_id)
    try:
        folder = get_object_or_404(Folder, id=folder_id)
        category_id = request.POST.get('category_id')
        if not category_id:
            messages.error(request, 'Kategoriya tanlanmagan!')
            return redirect('folder_categories_manage', folder_id=folder_id)
        category = get_object_or_404(Category, id=category_id)
        if not FolderCategory.objects.filter(folder=folder, category=category).exists():
            FolderCategory.objects.create(folder=folder, category=category)
            messages.success(request, f'✅ "{category.name}" kategoriyasi papkaga qo\'shildi!')
        else:
            messages.warning(request, f'"{category.name}" allaqachon papkada mavjud!')
    except Exception as e:
        messages.error(request, f'Xatolik: {str(e)}')
    return redirect('folder_categories_manage', folder_id=folder_id)


@login_required
@user_passes_test(is_admin_user)
def folder_category_remove(request, folder_category_id):
    if request.method != 'POST':
        messages.error(request, 'Faqat POST so\'rov!')
        return redirect('admin_panel')
    try:
        fc = get_object_or_404(FolderCategory, id=folder_category_id)
        folder = fc.folder
        category_name = fc.category.name
        fc.delete()
        messages.success(request, f'🗑️ "{category_name}" kategoriyasi papkadan olib tashlandi!')
    except Exception as e:
        messages.error(request, f'Xatolik: {str(e)}')
        return redirect('admin_panel')
    return redirect('folder_categories_manage', folder_id=folder.id)


@login_required
@user_passes_test(is_admin_user)
def group_folders_manage(request, group_id):
    group = get_object_or_404(Group, id=group_id)
    assigned_folders = GroupFolder.objects.filter(group=group).select_related('folder')
    assigned_ids = [gf.folder.id for gf in assigned_folders]

    assigned_data = []
    for gf in assigned_folders:
        try:
            config = FolderGroupConfig.objects.get(folder=gf.folder, group=group)
        except FolderGroupConfig.DoesNotExist:
            config = FolderGroupConfig.objects.create(
                folder=gf.folder, group=group,
                categories_to_select=1, is_active=True
            )
        folder_cats = FolderCategory.objects.filter(folder=gf.folder).select_related('category')
        assigned_data.append((gf, config, folder_cats))

    available_folders = Folder.objects.exclude(id__in=assigned_ids)
    all_categories = Category.objects.all().order_by('name')
    return render(request, 'groups/group_folders_manage.html', {
        'group': group,
        'assigned_data': assigned_data,
        'available_folders': available_folders,
        'all_categories': all_categories,
    })


@login_required
@user_passes_test(is_admin_user)
def group_folder_add(request, group_id):
    if request.method != 'POST':
        messages.error(request, 'Faqat POST so\'rov!')
        return redirect('group_folders_manage', group_id=group_id)
    try:
        group = get_object_or_404(Group, id=group_id)
        folder_id = request.POST.get('folder_id')
        if not folder_id:
            messages.error(request, 'Papka tanlanmagan!')
            return redirect('group_folders_manage', group_id=group_id)
        folder = get_object_or_404(Folder, id=folder_id)
        if not GroupFolder.objects.filter(group=group, folder=folder).exists():
            GroupFolder.objects.create(group=group, folder=folder)
            FolderGroupConfig.objects.get_or_create(
                folder=folder, group=group,
                defaults={'categories_to_select': 1, 'is_active': True}
            )
            messages.success(request, f'✅ "{folder.name}" papkasi guruhga qo\'shildi!')
        else:
            messages.warning(request, f'"{folder.name}" allaqachon guruhga biriktirilgan!')
    except Exception as e:
        messages.error(request, f'Xatolik: {str(e)}')
    return redirect('group_folders_manage', group_id=group_id)


@login_required
@user_passes_test(is_admin_user)
def group_folder_remove(request, group_folder_id):
    if request.method != 'POST':
        messages.error(request, 'Faqat POST so\'rov!')
        return redirect('admin_panel')
    try:
        gf = get_object_or_404(GroupFolder, id=group_folder_id)
        group = gf.group
        folder_name = gf.folder.name
        gf.delete()
        messages.success(request, f'🗑️ "{folder_name}" papkasi guruhdan olib tashlandi!')
    except Exception as e:
        messages.error(request, f'Xatolik: {str(e)}')
        return redirect('admin_panel')
    return redirect('group_folders_manage', group_id=group.id)


@login_required
@user_passes_test(is_admin_user)
@csrf_exempt
def folder_group_config_edit_api(request, config_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Faqat POST so\'rov'})
    try:
        data = json.loads(request.body)
        config = get_object_or_404(FolderGroupConfig, id=config_id)
        config.categories_to_select = data.get('categories_to_select', 1)
        config.randomize_categories = data.get('randomize_categories', True)
        config.is_active = data.get('is_active', True)
        config.save()
        return JsonResponse({'success': True, 'message': 'Sozlama tahrirlandi'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# ==================== END FOLDER VIEWS ====================


@login_required
@user_passes_test(is_admin_user)
def rules_edit(request):
    rules, created = Rules.objects.get_or_create(id=1)
    if request.method == 'POST':
        try:
            rules.video_url = request.POST.get('video_url', '')
            if request.FILES.get('video_file'):
                rules.video_file = request.FILES['video_file']
            if request.FILES.get('image1'):
                rules.image1 = request.FILES['image1']
            rules.image1_title = request.POST.get('image1_title', 'Imtihon tartibi')
            rules.image1_description = request.POST.get('image1_description', '')
            if request.FILES.get('image2'):
                rules.image2 = request.FILES['image2']
            rules.image2_title = request.POST.get('image2_title', 'Baholash mezonlari')
            rules.image2_description = request.POST.get('image2_description', '')
            rules.rules_text = request.POST.get('rules_text', '')
            rules.save()
            messages.success(request, '✅ Qonun va qoidalar saqlandi!')
            return redirect('rules_edit')
        except Exception as e:
            messages.error(request, f'Xatolik: {str(e)}')
    return render(request, 'groups/rules_edit.html', {'rules': rules})


@login_required
@user_passes_test(is_admin_user)
@csrf_exempt
def get_group_api(request, group_id):
    try:
        group = Group.objects.get(id=group_id)
        return JsonResponse({'success': True, 'id': group.id, 'name': group.name, 'teacher': group.teacher})
    except Group.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Guruh topilmadi'}, status=404)


@login_required
@user_passes_test(is_admin_user)
@csrf_exempt
def category_group_config_edit_api(request, config_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Faqat POST so\'rov'})
    try:
        data = json.loads(request.body)
        config = get_object_or_404(CategoryGroupConfig, id=config_id)
        config.questions_count = data.get('questions_count', 3)
        config.random_order = data.get('random_order', True)
        config.is_active = data.get('is_active', True)
        config.save()
        return JsonResponse({'success': True, 'message': 'Sozlama tahrirlandi'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@user_passes_test(is_admin_user)
@csrf_exempt
def category_group_config_delete_api(request, config_id):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Faqat POST so\'rov'})
    try:
        config = get_object_or_404(CategoryGroupConfig, id=config_id)
        config.delete()
        return JsonResponse({'success': True, 'message': 'Sozlama o\'chirildi'})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@csrf_exempt
@require_http_methods(["GET", "POST"])
def check_exam_api(request):
    try:
        student = request.user.student_profile if hasattr(request.user, 'student_profile') else None
        if not student:
            return JsonResponse({'success': False, 'error': 'Student profile not found'}, status=400)

        active_session = ExamSession.objects.filter(group=student.group, is_active=True).first()
        if active_session:
            return JsonResponse({'success': True, 'has_active_exam': True, 'exam_id': active_session.id})
        return JsonResponse({'success': True, 'has_active_exam': False})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)






@login_required
@user_passes_test(is_admin_user)
def student_attempts_api(request, student_id):
    try:
        student = Student.objects.get(id=student_id)
        
        # MUHIM: Distinct natijalarni olish (dublikatlarsiz)
        # Har bir attempt_number uchun faqat eng oxirgi natijani olamiz
        from django.db.models import Max, Subquery, OuterRef
        
        # Har bir attempt_number uchun eng oxirgi natijaning ID'sini topish
        latest_per_attempt = QuizResult.objects.filter(
            student=student
        ).values('attempt_number').annotate(
            latest_id=Max('id')
        ).values_list('latest_id', flat=True)
        
        # Faqat eng oxirgi natijalarni olish
        results = QuizResult.objects.filter(
            id__in=latest_per_attempt
        ).order_by('-attempt_number')
        
        attempts = []
        best_score = 0
        total_score = 0
        
        for r in results:
            # Har bir urinish uchun ballni hisoblash
            total_ball = float(r.total_questions) if r.total_questions else 0
            score_value = round((float(r.score) / 100) * total_ball, 1) if total_ball > 0 else 0
            attempts.append({
                'id': r.id,
                'attempt_number': r.attempt_number,
                'score': score_value,
                'total': total_ball,
                'total_ball': total_ball,
                'percentage': float(r.percentage),
                'submitted_at': timezone.localtime(r.submitted_at).strftime('%Y-%m-%d %H:%M:%S'),
                'is_last': (results.first().id == r.id) if results.exists() else False
            })
            
            if score_value > best_score:
                best_score = score_value
            total_score += score_value
        
        avg_score = round(total_score / len(results), 1) if results else 0
        
        # Debug uchun
        print(f"Student {student.full_name} - Found {len(attempts)} attempts (unique)")
        for a in attempts:
            print(f"  Attempt #{a['attempt_number']}: score={a['score']}, id={a['id']}")
        
        return JsonResponse({
            'success': True,
            'student_name': student.full_name,
            'total_attempts': len(attempts),
            'best_score': best_score,
            'avg_score': avg_score,
            'attempts': attempts
        })
        
    except Student.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Student topilmadi'})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': str(e)})



from django.shortcuts import render

def offline_view(request):
    return render(request, 'groups/offline.html')






@staff_member_required
def admin_question_list(request):
    questions = QuizQuestion.objects.all().select_related('category').order_by('-created_at')
    search = request.GET.get('search', '')
    if search:
        questions = questions.filter(
            Q(question_text__icontains=search) |
            Q(correct_answer__icontains=search) |
            Q(correct_sentence__icontains=search)
        )
    question_type = request.GET.get('type', '')
    if question_type:
        questions = questions.filter(question_type=question_type)
    category_id = request.GET.get('category', '')
    if category_id:
        questions = questions.filter(category_id=category_id)

    return render(request, 'groups/admin_question_list.html', {
        'questions': questions,
        'categories': Category.objects.all(),
        'search': search,
        'selected_type': question_type,
        'selected_category': category_id,
    })

@staff_member_required
def admin_question_add(request):
    if request.method == 'POST':
        question_type = request.POST.get('question_type')
        category_id = request.POST.get('category')

        if not question_type:
            messages.error(request, "Iltimos, savol turini tanlang!")
            return redirect('admin_question_add')

        try:
            category = Category.objects.get(id=category_id)
            points = int(request.POST.get('points', 1))

            audio_updated = False

            # Audio fayl yuklash
            if 'category_audio' in request.FILES:
                audio_file = request.FILES['category_audio']
                if audio_file.size <= 10 * 1024 * 1024:
                    if category.audio_file:
                        try:
                            category.audio_file.delete(save=False)
                        except Exception:
                            pass
                    category.audio_file = audio_file
                    audio_updated = True
                    messages.success(request, "✅ Audio fayl yuklandi!")
                else:
                    messages.warning(request, 'Audio fayl 10MB dan kichik bo\'lishi kerak!')

            # Audio faylni o'chirish
            if request.POST.get('clear_category_audio') == 'true':
                if category.audio_file:
                    try:
                        category.audio_file.delete(save=False)
                    except Exception:
                        pass
                    category.audio_file = None
                    audio_updated = True
                    messages.success(request, "🗑️ Audio fayl o'chirildi!")

            # Maksimal eshitish soni
            max_plays = request.POST.get('category_max_audio_plays')
            if max_plays:
                try:
                    category.max_audio_plays = int(max_plays)
                    audio_updated = True
                except ValueError:
                    pass

            # Audio ko'rsatmasi
            instruction = request.POST.get('category_audio_instruction')
            if instruction is not None:
                category.audio_instruction = instruction
                audio_updated = True

            if audio_updated:
                category.save()

            # ============ 1. FILL BLANK (Word Bank) ============
            if question_type == 'fill_blank':
                question_text = request.POST.get('fb_question_text', '').strip()
                correct_answer = request.POST.get('fb_correct_answer', '').strip()
                if not question_text or not correct_answer:
                    messages.error(request, "Savol matni va to'g'ri javobni kiriting!")
                else:
                    QuizQuestion.objects.create(
                        category=category, question_type='fill_blank',
                        question_text=question_text, correct_answer=correct_answer,
                        points=points
                    )
                    messages.success(request, "✅ Bo'sh joy (Word Bank) savoli qo'shildi!")
                    return redirect('admin_question_list')

            # ============ 2. FILL BLANK NO WORD ============
            elif question_type == 'fill_blank_no_word':
                question_text = request.POST.get('fbnw_question_text', '').strip()
                correct_answer = request.POST.get('fbnw_correct_answer', '').strip()
                if not question_text or not correct_answer:
                    messages.error(request, "Savol matni va to'g'ri javobni kiriting!")
                else:
                    QuizQuestion.objects.create(
                        category=category, question_type='fill_blank_no_word',
                        question_text=question_text, correct_answer=correct_answer,
                        points=points
                    )
                    messages.success(request, "✅ Bo'sh joy (variantlarsiz) savoli qo'shildi!")
                    return redirect('admin_question_list')

            # ============ 3. SENTENCE ARRANGEMENT ============
            elif question_type == 'sentence_arrangement':
                scrambled_words = request.POST.get('sa_scrambled_words', '').strip()
                correct_sentence = request.POST.get('sa_correct_sentence', '').strip()
                if not scrambled_words or not correct_sentence:
                    messages.error(request, "Barcha maydonlarni to'ldiring!")
                else:
                    QuizQuestion.objects.create(
                        category=category, question_type='sentence_arrangement',
                        scrambled_words=scrambled_words, correct_sentence=correct_sentence,
                        points=points
                    )
                    messages.success(request, "✅ So'z tartibi savoli qo'shildi!")
                    return redirect('admin_question_list')

            # ============ 4. READING COMPREHENSION ============
            elif question_type == 'reading_comprehension':
                reading_mode = request.POST.get('reading_mode', 'new')
                if reading_mode == 'new':
                    reading_title = request.POST.get('reading_title', '').strip()
                    reading_content = request.POST.get('reading_content', '').strip()
                    if not reading_title or not reading_content:
                        messages.error(request, "Matn sarlavhasi va mazmunini kiriting!")
                        return redirect('admin_question_add')
                    reading_text_obj = ReadingText.objects.create(
                        title=reading_title, content=reading_content, category=category
                    )
                else:
                    existing_text_id = request.POST.get('existing_reading_text')
                    if not existing_text_id:
                        messages.error(request, "Matn tanlanmagan!")
                        return redirect('admin_question_add')
                    reading_text_obj = ReadingText.objects.get(id=existing_text_id)

                questions_saved = 0
                for key, value in request.POST.items():
                    if key.startswith('reading_question_'):
                        idx = key.split('_')[-1]
                        q_text = request.POST.get(f'reading_question_{idx}', '').strip()
                        q_answer = request.POST.get(f'reading_answer_{idx}', '').strip()
                        if q_text and q_answer:
                            ReadingQuestion.objects.create(
                                reading_text=reading_text_obj,
                                question_text=q_text,
                                correct_answer=q_answer,
                                order=int(idx) if idx.isdigit() else 0
                            )
                            questions_saved += 1

                if questions_saved == 0:
                    messages.warning(request, "Kamida bitta savol kiriting!")
                    return redirect('admin_question_add')

                QuizQuestion.objects.create(
                    category=category, question_type='reading_comprehension',
                    reading_text=reading_text_obj,
                    question_text=f"📖 {reading_text_obj.title}", correct_answer="",
                    points=points
                )
                messages.success(request, f"✅ Matnli savol qo'shildi! ({questions_saved} ta savol)")
                return redirect('admin_question_list')

            # ============ 5. TRUE/FALSE ============
            elif question_type == 'true_false':
                question_text = request.POST.get('tf_question_text', '').strip()
                correct_answer = request.POST.get('tf_correct_answer', '').strip()
                if not question_text:
                    messages.error(request, "Savol matnini kiriting!")
                elif not correct_answer:
                    messages.error(request, "To'g'ri javobni tanlang!")
                else:
                    QuizQuestion.objects.create(
                        category=category, question_type='true_false',
                        question_text=question_text, correct_answer=correct_answer,
                        points=points
                    )
                    messages.success(request, "✅ To'g'ri/Noto'g'ri savoli qo'shildi!")
                    return redirect('admin_question_list')

            # ============ 6. MULTIPLE CHOICE ============
            elif question_type == 'multiple_choice':
                question_text = request.POST.get('mc_question_text', '').strip()
                options = []
                for key, value in request.POST.items():
                    if key.startswith('option_') and value.strip():
                        options.append(value.strip())
                correct_option_index = request.POST.get('correct_option')
                if correct_option_index and correct_option_index.isdigit():
                    correct_option_index = int(correct_option_index)
                else:
                    correct_option_index = None

                if not question_text:
                    messages.error(request, "Savol matnini kiriting!")
                elif len(options) < 2:
                    messages.error(request, "Kamida 2 ta variant kiriting!")
                elif correct_option_index is None or correct_option_index >= len(options):
                    messages.error(request, "To'g'ri variantni tanlang!")
                else:
                    correct_answer = options[correct_option_index]
                    QuizQuestion.objects.create(
                        category=category, question_type='multiple_choice',
                        question_text=question_text,
                        correct_answer=correct_answer,
                        scrambled_words=json.dumps(options),
                        points=points
                    )
                    messages.success(request, "✅ Test varianti savoli qo'shildi!")
                    return redirect('admin_question_list')

            # ============ 7. UNDERLINE CORRECT (TO'G'RI SO'Z) - TO'G'RILANGAN ============
            elif question_type == 'underline_correct':
                sentence_text = request.POST.get('uc_sentence_text', '').strip()
                correct_answer = request.POST.get('uc_correct_answer', '').strip()
                
                if not sentence_text:
                    messages.error(request, "Gap matnini kiriting!")
                elif not correct_answer:
                    messages.error(request, "To'g'ri javobni kiriting!")
                else:
                    # Matndan variantlarni to'g'ri ajratish
                    options = []
                    
                    if '/' in sentence_text:
                        slash_index = sentence_text.find('/')
                        
                        # Slash dan oldingi qismdagi oxirgi so'zni olish
                        before_slash = sentence_text[:slash_index].strip()
                        before_words = before_slash.split()
                        left_option = before_words[-1] if before_words else ''
                        
                        # Slash dan keyingi qismdagi birinchi so'zni olish
                        after_slash = sentence_text[slash_index + 1:].strip()
                        after_words = after_slash.split()
                        right_option = after_words[0] if after_words else ''
                        
                        # Faqat ikkala variantni qo'shish
                        if left_option:
                            options.append(left_option)
                        if right_option:
                            options.append(right_option)
                    else:
                        # Agar / bo'lmasa, butun matnni bitta variant sifatida saqlash
                        options = [sentence_text]
                    
                    # Hech qanday variant topilmasa
                    if not options or len(options) < 2:
                        messages.error(request, "Matn ichida / belgisi bilan ajratilgan 2 ta variant topilmadi! Misol: 'in / by'")
                        return redirect('admin_question_add')
                    
                    # To'g'ri javob variantlardan biriga tengligini tekshirish
                    if correct_answer not in options:
                        messages.error(request, f"To'g'ri javob '{correct_answer}' variantlar ichida emas! Variantlar: {', '.join(options)}")
                        return redirect('admin_question_add')
                    
                    # scrambled_words ga variantlarni JSON sifatida saqlash
                    scrambled_words_json = json.dumps(options)
                    
                    QuizQuestion.objects.create(
                        category=category, 
                        question_type='underline_correct',
                        question_text=sentence_text, 
                        correct_answer=correct_answer,
                        scrambled_words=scrambled_words_json,
                        points=points
                    )
                    messages.success(request, "✅ To'g'ri so'zni tanlash savoli qo'shildi!")
                    return redirect('admin_question_list')

            # ============ 8. MATCHING (MOSLASHTIRISH) ============
            elif question_type == 'matching':
                instruction = request.POST.get('matching_instruction', '').strip()
                left_items = []
                for key, value in request.POST.items():
                    if key.startswith('left_item_') and value.strip():
                        left_items.append(value.strip())
                right_items = []
                for key, value in request.POST.items():
                    if key.startswith('right_item_') and value.strip():
                        right_items.append(value.strip())
                matches = {}
                for key, value in request.POST.items():
                    if key.startswith('match_answer_') and value:
                        idx = int(key.split('_')[-1])
                        matches[str(idx + 1)] = value

                if len(left_items) < 2:
                    messages.error(request, "Kamida 2 ta chap tomon elementi kiriting!")
                elif len(right_items) < 2:
                    messages.error(request, "Kamida 2 ta o'ng tomon elementi kiriting!")
                elif len(matches) != len(left_items):
                    messages.error(request, f"Barcha {len(left_items)} ta elementga javob tanlang!")
                else:
                    matching_data = {'left': left_items, 'right': right_items}
                    QuizQuestion.objects.create(
                        category=category, question_type='matching',
                        question_text=instruction,
                        correct_answer=json.dumps(matches),
                        scrambled_words=json.dumps(matching_data),
                        points=points
                    )
                    messages.success(request, "✅ Moslashtirish savoli qo'shildi!")
                    return redirect('admin_question_list')

            # ============ 9. CLOZE MULTIPLE BLANKS ============
            elif question_type == 'cloze_multiple_blanks':
                cloze_text = request.POST.get('cloze_question_text', '').strip()
                blank_options_json = request.POST.get('cloze_blank_options', '{}')
                correct_answers_json = request.POST.get('cloze_correct_answers', '{}')

                try:
                    blank_options = json.loads(blank_options_json)
                    correct_answers = json.loads(correct_answers_json)
                except json.JSONDecodeError:
                    blank_options = {}
                    correct_answers = {}

                if not cloze_text:
                    messages.error(request, "Matnni kiriting!")
                elif not blank_options or len(blank_options) == 0:
                    messages.error(request, "Hech qanday bo'sh joy aniqlanmadi!")
                else:
                    QuizQuestion.objects.create(
                        category=category, question_type='cloze_multiple_blanks',
                        question_text=cloze_text,
                        correct_answer=json.dumps(correct_answers),
                        blank_options=blank_options,
                        scrambled_words=json.dumps(blank_options),
                        points=points
                    )
                    messages.success(request, f"✅ Ko'p bo'sh joy savoli qo'shildi! ({len(blank_options)} ta)")
                    return redirect('admin_question_list')

            # ============ 10. COMPLETE THE WORDS ============
            elif question_type == 'complete_the_words':
                sentence_text = request.POST.get('ctw_sentence_text', '').strip()
                correct_answer = request.POST.get('ctw_correct_answer', '').strip().lower()
                if not sentence_text:
                    messages.error(request, "Gap matnini kiriting!")
                elif not correct_answer:
                    messages.error(request, "To'g'ri javobni kiriting!")
                else:
                    QuizQuestion.objects.create(
                        category=category, question_type='complete_the_words',
                        question_text=sentence_text, correct_answer=correct_answer,
                        points=points
                    )
                    messages.success(request, "✅ So'zlarni to'ldirish savoli qo'shildi!")
                    return redirect('admin_question_list')

            elif question_type == 'writing':
                topic = request.POST.get('w_topic', '').strip()
                if not topic:
                    messages.error(request, "Mavzu matnini kiriting!")
                else:
                    QuizQuestion.objects.create(
                        category=category, question_type='writing',
                        question_text=topic, correct_answer='',
                        points=points
                    )
                    messages.success(request, "✅ Yozma ish (Writing) savoli qo'shildi!")
                    return redirect('admin_question_list')

            elif question_type == 'speaking':
                topic = request.POST.get('s_topic', '').strip()
                if not topic:
                    messages.error(request, "Mavzu matnini kiriting!")
                else:
                    QuizQuestion.objects.create(
                        category=category, question_type='speaking',
                        question_text=topic, correct_answer='',
                        points=points
                    )
                    messages.success(request, "✅ Og'zaki (Speaking) savoli qo'shildi!")
                    return redirect('admin_question_list')

            else:
                messages.error(request, f"Noto'g'ri savol turi: '{question_type}'")

        except Category.DoesNotExist:
            messages.error(request, "Kategoriya topilmadi!")
        except Exception as e:
            import traceback
            traceback.print_exc()
            messages.error(request, f'Xatolik: {str(e)}')

    categories = Category.objects.all()
    reading_texts = ReadingText.objects.all().select_related('category')
    return render(request, 'groups/admin_question_add.html', {
        'categories': categories,
        'reading_texts': reading_texts,
    })

@staff_member_required
def admin_question_edit(request, pk):
    question = get_object_or_404(QuizQuestion, id=pk)
    if request.method == 'POST':
        question_type = request.POST.get('question_type')
        category_id = request.POST.get('category')
        try:
            category = Category.objects.get(id=category_id)
            question.category = category
            question.question_type = question_type
            if question_type == 'fill_blank':
                question.question_text = request.POST.get('question_text')
                question.correct_answer = request.POST.get('correct_answer')
            else:
                question.scrambled_words = request.POST.get('scrambled_words')
                question.correct_sentence = request.POST.get('correct_sentence')
            question.save()
            messages.success(request, "Savol tahrirlandi!")
            return redirect('admin_question_list')
        except Category.DoesNotExist:
            messages.error(request, "Kategoriya topilmadi!")
    return render(request, 'groups/admin_question_edit.html', {
        'question': question,
        'categories': Category.objects.all(),
    })


@staff_member_required
def admin_question_delete(request, pk):
    question = get_object_or_404(QuizQuestion, id=pk)
    if request.method == 'POST':
        question.delete()
        messages.success(request, "Savol o'chirildi!")
        return redirect('admin_question_list')
    return render(request, 'groups/admin_question_confirm_delete.html', {'question': question})


@login_required
@user_passes_test(is_admin_user)
def reading_texts_list(request):
    texts = ReadingText.objects.all().select_related('category').annotate(
        questions_count=Count('reading_questions')
    )
    return render(request, 'groups/reading_texts_list.html', {
        'texts': texts,
        'total_texts': texts.count(),
    })


@login_required
@user_passes_test(is_admin_user)
def reading_text_edit(request, pk):
    text = get_object_or_404(ReadingText, id=pk)
    if request.method == 'POST':
        text.title = request.POST.get('title', '').strip()
        text.content = request.POST.get('content', '').strip()
        category_id = request.POST.get('category')
        if category_id:
            text.category_id = category_id
        text.save()

        for key, value in request.POST.items():
            if key.startswith('question_text_'):
                q_id = key.replace('question_text_', '')
                question_text = request.POST.get(f'question_text_{q_id}', '')
                correct_answer = request.POST.get(f'correct_answer_{q_id}', '')
                if question_text and correct_answer:
                    if q_id.isdigit():
                        ReadingQuestion.objects.filter(id=q_id).update(
                            question_text=question_text, correct_answer=correct_answer
                        )
                    else:
                        ReadingQuestion.objects.create(
                            reading_text=text,
                            question_text=question_text,
                            correct_answer=correct_answer
                        )

        messages.success(request, 'Matn va savollar yangilandi!')
        return redirect('reading_texts_list')

    return render(request, 'groups/reading_text_edit.html', {
        'text': text,
        'categories': Category.objects.all(),
        'questions': text.reading_questions.all().order_by('order'),
    })


# Eski URL'lar uchun (slash boshida)
@login_required
@csrf_exempt
def check_audio_play_old(request):
    return check_audio_play_api(request)


@login_required
@csrf_exempt
def record_audio_play_old(request):
    return record_audio_play_api(request)








# models.py ga qo'shimcha maydon (agar bo'lmasa)
# ExamControl modeliga is_paused qo'shing:





@login_required
@user_passes_test(is_admin_user)
@csrf_exempt
@require_http_methods(["POST"])
def pause_exam_api(request):
    """Testni vaqtincha to'xtatish (pauza)"""
    try:
        data = json.loads(request.body)
        group_id = data.get('group_id')
        if not group_id:
            return JsonResponse({'success': False, 'message': 'group_id kerak'})

        group = Group.objects.get(id=group_id)
        config, _ = GroupExamConfig.objects.get_or_create(group=group)
        now = timezone.now()
        
        exam_control, _ = ExamControl.objects.get_or_create(group=group)
        
        # Agar test aktiv bo'lmasa
        if not exam_control.is_active:
            return JsonResponse({'success': False, 'message': 'Test aktiv emas!'})
        
        # O'tgan vaqtni hisoblash (started_at dan hozirgacha)
        if exam_control.started_at:
            elapsed = (now - exam_control.started_at).total_seconds()
            exam_control.elapsed_time = int(elapsed)
        else:
            exam_control.elapsed_time = 0
        
        exam_control.is_active = False
        exam_control.is_paused = True
        exam_control.paused_at = now
        exam_control.save()
        
        # Cache ni yangilash
        cache.set(f'exam_active_{group_id}', False, timeout=86400)
        cache.set(f'exam_paused_{group_id}', True, timeout=86400)
        cache.set(f'exam_elapsed_time_{group_id}', exam_control.elapsed_time, timeout=86400)
        
        # QuizSession ni pauzaga o'tkazish
        quiz_session = QuizSession.objects.filter(group=group, is_active=True).first()
        if quiz_session:
            quiz_session.is_active = False
            quiz_session.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Test pauzaga qo\'yildi!',
            'is_active': False,
            'is_paused': True,
            'elapsed_time': exam_control.elapsed_time
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def quiz_check_status(request):
    """Test holatini tekshirish - studentlar uchun"""
    try:
        data = json.loads(request.body)
        group_id = data.get('group_id')
        if not group_id:
            return JsonResponse({'success': False, 'error': 'group_id kerak'})

        group = Group.objects.get(id=group_id)
        config, _ = GroupExamConfig.objects.get_or_create(group=group)
        
        # Cache dan tezroq olish
        is_active = cache.get(f'exam_active_{group_id}')
        is_paused = cache.get(f'exam_paused_{group_id}')
        start_time_str = cache.get(f'exam_start_time_{group_id}')
        
        # Agar cache da bo'lmasa, database dan olish
        if is_active is None:
            try:
                exam_control = ExamControl.objects.get(group_id=group_id)
                is_active = exam_control.is_active
                is_paused = exam_control.is_paused
                start_time_str = exam_control.started_at.isoformat() if exam_control.started_at else None
                elapsed_time = exam_control.elapsed_time
            except ExamControl.DoesNotExist:
                is_active = False
                is_paused = False
                elapsed_time = 0
        else:
            # Cache da bo'lsa, elapsed_time ni ham olish
            elapsed_time = cache.get(f'exam_elapsed_time_{group_id}', 0)
        
        remaining_time = None
        total_time = config.time_limit * 60 if config.time_limit > 0 else 0
        
        if is_active and start_time_str and total_time > 0:
            # Aktiv test - o'tgan vaqtni hisoblash
            try:
                start_time = datetime.fromisoformat(start_time_str)
                if timezone.is_naive(start_time):
                    start_time = timezone.make_aware(start_time)
                now = timezone.now()
                elapsed = (now - start_time).total_seconds()
                remaining_time = max(0, int(total_time - elapsed))
            except:
                remaining_time = total_time
                
        elif is_paused and total_time > 0:
            # Pauza holati
            remaining_time = max(0, total_time - elapsed_time)
            
        elif total_time > 0:
            remaining_time = total_time
        
        # VAQT TUGASA - testni avtomatik to'xtatish
        if remaining_time is not None and remaining_time <= 0 and is_active and not is_paused:
            try:
                exam_control = ExamControl.objects.get(group_id=group_id)
                exam_control.is_active = False
                exam_control.is_paused = False
                exam_control.save()
                cache.set(f'exam_active_{group_id}', False, 300)
                cache.set(f'exam_paused_{group_id}', False, 300)
                is_active = False
                is_paused = False
            except ExamControl.DoesNotExist:
                pass

        return JsonResponse({
            'success': True,
            'is_active': is_active,
            'is_paused': is_paused,
            'remaining_time': remaining_time,
            'elapsed_time': elapsed_time
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e), 'is_active': False, 'is_paused': False}, status=500)

























def format_time(seconds):
    """Vaqtni formatlash"""
    minutes = seconds // 60
    secs = seconds % 60
    return f"{minutes}:{secs:02d}"

@login_required
@user_passes_test(is_admin_user)
@csrf_exempt
@require_http_methods(["POST"])
def resume_exam_api(request):
    """Pauzadagi testni qayta boshlash - vaqtni to'g'ri hisoblash"""
    try:
        data = json.loads(request.body)
        group_id = data.get('group_id')
        if not group_id:
            return JsonResponse({'success': False, 'message': 'group_id kerak'})

        group = Group.objects.get(id=group_id)
        config, _ = GroupExamConfig.objects.get_or_create(group=group)
        now = timezone.now()
        
        exam_control, _ = ExamControl.objects.get_or_create(group=group)
        
        # Agar test pauzada bo'lmasa
        if not exam_control.is_paused:
            return JsonResponse({'success': False, 'message': 'Test pauzada emas!'})
        
        total_time = config.time_limit * 60  # Jami vaqt (sekund)
        elapsed_time = exam_control.elapsed_time  # O'tgan vaqt (sekund)
        remaining_time = max(0, total_time - elapsed_time)  # Qolgan vaqt
        
        # MUHIM: Yangi start vaqtini hisoblash - o'tgan vaqtni hisobga olgan holda
        # Test boshlangan vaqt = Hozirgi vaqt - o'tgan vaqt
        new_started_at = now - timezone.timedelta(seconds=elapsed_time)
        
        exam_control.is_active = True
        exam_control.is_paused = False
        exam_control.started_at = new_started_at
        exam_control.paused_at = None
        # elapsed_time ni o'zgartirmaymiz! (o'tgan vaqt saqlanib qoladi)
        exam_control.save()
        
        # Cache ni yangilash
        cache.set(f'exam_active_{group_id}', True, timeout=86400)
        cache.delete(f'exam_paused_{group_id}')
        cache.set(f'exam_start_time_{group_id}', new_started_at.isoformat(), timeout=86400)
        
        # Tugash vaqtini hisoblash
        if config.time_limit > 0:
            end_time = new_started_at + timezone.timedelta(seconds=total_time)
            cache.set(f'exam_end_time_{group_id}', end_time.isoformat(), timeout=86400)
        
        # QuizSession ni qayta aktivlashtirish
        quiz_session = QuizSession.objects.filter(group=group).order_by('-started_at').first()
        if quiz_session:
            quiz_session.is_active = True
            quiz_session.save()
        else:
            quiz_session = QuizSession.objects.create(
                group=group, is_active=True, started_at=now, created_by=request.user
            )
        
        # Qolgan vaqtni formatlash
        remaining_min = remaining_time // 60
        remaining_sec = remaining_time % 60
        
        return JsonResponse({
            'success': True,
            'message': f'Test qayta boshlandi! Qolgan vaqt: {remaining_min}:{remaining_sec:02d}',
            'is_active': True,
            'is_paused': False,
            'remaining_time': remaining_time,
            'elapsed_time': elapsed_time
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@csrf_exempt
@require_http_methods(["POST"])
def get_remaining_time_api(request):
    """Qolgan vaqtni olish (student uchun)"""
    try:
        data = json.loads(request.body)
        group_id = data.get('group_id')
        if not group_id:
            return JsonResponse({'success': False, 'message': 'group_id kerak'})

        group = Group.objects.get(id=group_id)
        config, _ = GroupExamConfig.objects.get_or_create(group=group)
        
        if config.time_limit <= 0:
            return JsonResponse({'success': True, 'remaining_time': None})
        
        try:
            exam_control = ExamControl.objects.get(group_id=group_id)
            total_time = config.time_limit * 60
            
            if exam_control.is_paused:
                # Pauzada - elapsed_time dan hisoblash
                remaining_time = max(0, total_time - exam_control.elapsed_time)
            elif exam_control.is_active and exam_control.started_at:
                # Aktiv - started_at dan hisoblash
                now = timezone.now()
                elapsed = (now - exam_control.started_at).total_seconds()
                remaining_time = max(0, int(total_time - elapsed))
            else:
                remaining_time = total_time
                
        except ExamControl.DoesNotExist:
            remaining_time = config.time_limit * 60
        
        # VAQT TUGASA - testni avtomatik to'xtatish
        if remaining_time is not None and remaining_time <= 0 and 'exam_control' in locals() and exam_control.is_active and not exam_control.is_paused:
            try:
                exam_control.is_active = False
                exam_control.is_paused = False
                exam_control.save()
                cache.set(f'exam_active_{group_id}', False, 300)
                cache.set(f'exam_paused_{group_id}', False, 300)
            except Exception:
                pass

        return JsonResponse({
            'success': True,
            'remaining_time': remaining_time,
            'is_paused': exam_control.is_paused if 'exam_control' in locals() else False
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})





def check_matching_answer(question, answers, total_score, total_possible, question_results, qid):
    """Matching (Moslashtirish) savolini tekshirish va ball hisoblash"""
    correct_answers = question.get_matching_correct_answers()
    blanks_total = len(correct_answers)
    blanks_correct = 0
    blank_scores = {}
    pts_per_blank = round(question.points / blanks_total, 2)
    
    # Chap va o'ng tomon elementlarini olish
    left_items = question.get_matching_left_items()
    right_items = question.get_matching_right_items()
    
    # Harflarni matnga aylantirish uchun mapping
    letter_to_text = {}
    text_to_letter = {}
    for idx, item in enumerate(right_items):
        letter = chr(65 + idx)  # A, B, C, D...
        letter_to_text[letter] = item.strip().lower()
        letter_to_text[letter.lower()] = item.strip().lower()
        text_to_letter[item.strip().lower()] = letter
    
    for key, correct_val in correct_answers.items():
        # Student javobini topish
        user_val = ''
        
        # 1-usul: q_123_1 formatida
        field_name = f'q_{question.id}_{key}'
        if field_name in answers:
            user_val = answers[field_name]
        
        # 2-usul: q_123 formatida dict sifatida
        elif f'q_{question.id}' in answers and isinstance(answers[f'q_{question.id}'], dict):
            user_val = answers[f'q_{question.id}'].get(str(key), '')
            if not user_val:
                user_val = answers[f'q_{question.id}'].get(key, '')
        
        # 3-usul: to'g'ridan to'g'ri q_123_key formatida
        else:
            for ans_key, ans_val in answers.items():
                if str(question.id) in ans_key and str(key) in ans_key:
                    user_val = ans_val
                    break
        
        # Javobni tozalash
        if user_val:
            user_val = str(user_val).strip()
        
        correct_val_clean = str(correct_val).strip()
        
        # TO'G'RI JAVOBNI TEKSHIRISH
        is_correct = False
        
        # 1. To'g'ridan to'g'ri tenglik
        if user_val and user_val.lower() == correct_val_clean.lower():
            is_correct = True
        
        # 2. Student harf yozgan, to'g'ri javob harf
        elif user_val and len(user_val) == 1 and user_val.isalpha() and len(correct_val_clean) == 1 and correct_val_clean.isalpha():
            if user_val.upper() == correct_val_clean.upper():
                is_correct = True
        
        # 3. Student matn yozgan, to'g'ri javob harf
        elif user_val and len(correct_val_clean) == 1 and correct_val_clean.isalpha():
            expected_text = letter_to_text.get(correct_val_clean.upper(), '')
            if expected_text and user_val.lower() == expected_text:
                is_correct = True
        
        # 4. Student harf yozgan, to'g'ri javob matn
        elif user_val and len(user_val) == 1 and user_val.isalpha():
            expected_letter = text_to_letter.get(correct_val_clean.lower(), '')
            if expected_letter and user_val.upper() == expected_letter:
                is_correct = True
        
        # 5. Student matn yozgan, to'g'ri javob matn
        elif user_val and correct_val_clean:
            try:
                idx = int(key) - 1
                if 0 <= idx < len(left_items):
                    if len(correct_val_clean) == 1 and correct_val_clean.isalpha():
                        right_idx = ord(correct_val_clean.upper()) - 65
                        if 0 <= right_idx < len(right_items):
                            correct_text = right_items[right_idx].strip().lower()
                            if user_val.lower() == correct_text:
                                is_correct = True
                    else:
                        if user_val.lower() == correct_val_clean.lower():
                            is_correct = True
            except:
                pass
        
        # 6. To'g'ri javob variantlar ichida bo'lsa (| bilan ajratilgan)
        if not is_correct and user_val and '|' in correct_val_clean:
            variants = [v.strip().lower() for v in correct_val_clean.split('|')]
            if user_val.lower() in variants:
                is_correct = True
        
        # 7. SELECT dan kelgan value (masalan: "A. olma" formatida)
        if not is_correct and user_val and '.' in user_val:
            parts = user_val.split('.')
            if len(parts) == 2:
                letter_part = parts[0].strip()
                text_part = parts[1].strip()
                if letter_part.isalpha() and len(letter_part) == 1:
                    if letter_part.upper() == correct_val_clean.upper():
                        is_correct = True
                    elif text_part.lower() == correct_val_clean.lower():
                        is_correct = True
        
        if is_correct:
            blanks_correct += 1
            total_score += pts_per_blank
        
        blank_scores[key] = {
            'user_answer': user_val if user_val else 'Javob berilmagan',
            'is_correct': is_correct,
            'correct_answer': correct_val_clean
        }
    
    total_possible += question.points
    question_results[qid] = {
        'type': 'matching',
        'blanks_correct': blanks_correct,
        'blanks_total': blanks_total,
        'blanks': blank_scores
    }
    
    return total_score, total_possible, question_results




from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Q
import json

@login_required
def admin_writing_review(request, group_id):
    if not is_admin_user(request.user) and not is_teacher_user(request.user):
        messages.error(request, 'Sizda bu sahifani ko\'rish huquqi yo\'q!')
        return redirect('login')

    # Teacher faqat o'z guruhiga kirishi mumkin
    if is_teacher_user(request.user) and not is_admin_user(request.user):
        teacher = request.user.teacher_profile
        if not teacher.is_active:
            messages.error(request, 'Profilingiz faol emas!')
            return redirect('teacher_panel')
        group = get_object_or_404(Group, id=group_id)
        if not teacher.all_groups and not teacher.groups.filter(id=group.id).exists():
            messages.error(request, 'Siz bu guruhga biriktirilmagansiz!')
            return redirect('teacher_panel')
    else:
        group = get_object_or_404(Group, id=group_id)
    results = QuizResult.objects.filter(
        quiz_session__group=group
    ).order_by('-submitted_at')

    writing_entries = []
    seen_combinations = set()
    for result in results:
        if not result.answers:
            continue
        for qid_str, ans_data in result.answers.items():
            if isinstance(ans_data, dict) and ans_data.get('type') == 'writing':
                try:
                    question = QuizQuestion.objects.get(id=int(qid_str))
                except (QuizQuestion.DoesNotExist, ValueError):
                    continue
                combo = (result.student.id, question.id)
                if combo in seen_combinations:
                    continue
                writing_entries.append({
                    'result_id': result.id,
                    'question_id': question.id,
                    'student_name': result.student.full_name,
                    'topic': question.question_text,
                    'answer': ans_data.get('user_answer', ''),
                    'graded': ans_data.get('graded', False),
                    'earned_points': ans_data.get('earned_points', 0),
                    'max_points': question.points,
                    'submitted_at': result.submitted_at,
                })
                seen_combinations.add(combo)

    return render(request, 'groups/admin_writing_review.html', {
        'group': group,
        'writing_entries': writing_entries,
    })

@login_required
def admin_writing_grade_api(request, result_id, question_id):
    if not is_admin_user(request.user) and not is_teacher_user(request.user):
        return JsonResponse({'success': False, 'message': 'Huquq yo\'q'})

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST kerak'})
    try:
        data = json.loads(request.body)
        earned_points = float(data.get('earned_points', 0))
        result = get_object_or_404(QuizResult, id=result_id)
        qid_str = str(question_id)

        if qid_str not in result.answers or not isinstance(result.answers[qid_str], dict):
            return JsonResponse({'success': False, 'message': 'Bu savol topilmadi'})

        ans_data = result.answers[qid_str]
        if ans_data.get('type') != 'writing':
            return JsonResponse({'success': False, 'message': 'Bu writing savoli emas'})

        max_points = int(ans_data.get('max_points', 0))
        if max_points == 0:
            try:
                question = QuizQuestion.objects.get(id=question_id)
                max_points = question.points
            except QuizQuestion.DoesNotExist:
                return JsonResponse({'success': False, 'message': 'Savol topilmadi'})

        earned_points = max(0, min(earned_points, max_points))

        ans_data['earned_points'] = earned_points
        ans_data['graded'] = True
        ans_data['is_correct'] = earned_points > 0

        # Recalculate score (barcha savollar bo'yicha total_possible qayta hisoblanadi)
        total_score = 0
        total_possible = 0
        qids_in_answers = [int(k) for k in result.answers.keys() if str(k).isdigit()]
        questions = QuizQuestion.objects.filter(id__in=qids_in_answers) if qids_in_answers else QuizQuestion.objects.none()

        for q in questions:
            sqid = str(q.id)
            total_possible += q.points
            a = result.answers[sqid]
            if isinstance(a, dict):
                if a.get('type') == 'writing':
                    total_score += a.get('earned_points', 0)
                elif 'blanks' in a:
                    pts_per = round(q.points / max(a.get('blanks_total', 1), 1), 2)
                    total_score += a.get('blanks_correct', 0) * pts_per
                elif a.get('is_correct'):
                    total_score += q.points

        result.score = round((total_score / total_possible) * 100, 1) if total_possible > 0 else 0
        result.total_questions = total_possible
        result.answers[qid_str] = ans_data
        result.save()

        group = result.quiz_session.group
        if group and result.student:
            AssessmentScore.objects.update_or_create(
                student=result.student,
                group=group,
                assessment_type='written',
                defaults={
                    'score': int(earned_points),
                    'added_by': request.user,
                    'student_name_saved': result.student.full_name,
                    'group_name_saved': group.name,
                }
            )

        return JsonResponse({
            'success': True,
            'score': result.score,
            'total_possible': total_possible,
            'earned': earned_points,
            'total_score': round(total_score, 1),
        })

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@staff_member_required
def admin_question_list(request):
    questions = QuizQuestion.objects.all().select_related('category', 'reading_text').order_by('-created_at')
    search = request.GET.get('search', '')
    if search:
        questions = questions.filter(
            Q(question_text__icontains=search) |
            Q(correct_answer__icontains=search) |
            Q(correct_sentence__icontains=search)
        )
    question_type = request.GET.get('type', '')
    if question_type:
        questions = questions.filter(question_type=question_type)
    category_id = request.GET.get('category', '')
    if category_id:
        questions = questions.filter(category_id=category_id)

    return render(request, 'groups/admin_question_list.html', {
        'questions': questions,
        'categories': Category.objects.all(),
        'search': search,
        'selected_type': question_type,
        'selected_category': category_id,
    })


@staff_member_required
def admin_question_delete(request, pk):
    question = get_object_or_404(QuizQuestion, id=pk)
    if request.method == 'POST':
        try:
            question_title = str(question)
            question.delete()
            messages.success(request, f'✅ "{question_title}" savoli o\'chirildi!')
        except Exception as e:
            messages.error(request, f'Xatolik: {str(e)}')
        return redirect('admin_question_list')
    return redirect('admin_question_list')


from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse

@login_required
@user_passes_test(is_admin_user)
@csrf_exempt
def admin_question_delete_api(request, pk):
    """Savolni o'chirish API (AJAX uchun)"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Faqat POST so\'rov'})
    
    try:
        question = get_object_or_404(QuizQuestion, id=pk)
        question_title = str(question)
        question.delete()
        return JsonResponse({
            'success': True, 
            'message': f'✅ "{question_title}" savoli o\'chirildi!'
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


# ========================
# SERTIFIKAT VIEWLARI
# ========================

@login_required
@user_passes_test(is_admin_user)
def certificate_settings(request):
    setting = CertificateSetting.objects.first()
    certificates = Certificate.objects.all()[:50]

    if request.method == 'POST':
        threshold = request.POST.get('threshold_percentage', 50)
        bg_image = request.FILES.get('background_image')

        if not setting:
            setting = CertificateSetting.objects.create(
                threshold_percentage=threshold,
                is_active=True
            )
        else:
            setting.threshold_percentage = threshold

        if bg_image:
            setting.background_image = bg_image

        setting.save()
        messages.success(request, "Sertifikat sozlamalari saqlandi!")
        return redirect('certificate_settings')

    context = {
        'setting': setting,
        'certificates': certificates,
    }
    return render(request, 'groups/certificate_settings.html', context)


@login_required
@user_passes_test(is_admin_user)
def certificate_list(request):
    group_id = request.GET.get('group_id')
    certificates = Certificate.objects.filter(is_archived=False)

    if group_id:
        group = get_object_or_404(Group, id=group_id)
        certificates = certificates.filter(group_name=group.name)

    groups = Group.objects.all()

    if request.method == 'POST':
        cert_ids = request.POST.getlist('cert_ids')
        action = request.POST.get('action')
        count = len(cert_ids)
        if count > 0:
            certs = Certificate.objects.filter(id__in=cert_ids)
            if action == 'archive':
                certs.update(is_archived=True)
                messages.success(request, f"{count} ta sertifikat arxivlandi!")
            elif action == 'unarchive':
                certs.update(is_archived=False)
                messages.success(request, f"{count} ta sertifikat arxivdan chiqarildi!")
            elif action == 'delete':
                certs.delete()
                messages.success(request, f"{count} ta sertifikat o'chirildi!")
        return redirect(request.path + ('?group_id=' + group_id if group_id else ''))

    context = {
        'certificates': certificates,
        'groups': groups,
        'selected_group_id': int(group_id) if group_id else None,
    }
    return render(request, 'groups/certificate_list.html', context)


@login_required
@user_passes_test(is_admin_user)
def certificate_archive(request):
    group_id = request.GET.get('group_id')
    certificates = Certificate.objects.filter(is_archived=True)

    if group_id:
        group = get_object_or_404(Group, id=group_id)
        certificates = certificates.filter(group_name=group.name)

    groups = Group.objects.all()

    if request.method == 'POST':
        cert_ids = request.POST.getlist('cert_ids')
        action = request.POST.get('action')
        count = len(cert_ids)
        if count > 0:
            certs = Certificate.objects.filter(id__in=cert_ids)
            if action == 'unarchive':
                certs.update(is_archived=False)
                messages.success(request, f"{count} ta sertifikat arxivdan chiqarildi!")
            elif action == 'delete':
                certs.delete()
                messages.success(request, f"{count} ta sertifikat o'chirildi!")
        return redirect(request.path + ('?group_id=' + group_id if group_id else ''))

    context = {
        'certificates': certificates,
        'groups': groups,
        'selected_group_id': int(group_id) if group_id else None,
    }
    return render(request, 'groups/certificate_archive.html', context)


def generate_student_certificate(result):
    from .certificate_utils import save_certificate_pdf

    setting = CertificateSetting.objects.filter(is_active=True).first()
    if not setting or not setting.background_image:
        return None

    threshold = setting.threshold_percentage
    if result.score < threshold:
        return None

    existing = Certificate.objects.filter(quiz_result=result).first()
    if existing:
        return existing

    student_name = result.student_name_saved
    if not student_name and result.student:
        student_name = result.student.full_name
    group_name = result.group_name_saved

    teacher_name = None
    level = None
    group_obj = None
    if result.quiz_session and result.quiz_session.group_id:
        group_obj = result.quiz_session.group
    else:
        group_obj = Group.objects.filter(name=group_name).first()

    if group_obj:
        try:
            config = group_obj.exam_config
            if not config.certificate_enabled:
                return None
            teacher_name = config.certificate_teacher or group_obj.teacher
            level = config.certificate_level
        except:
            teacher_name = group_obj.teacher if group_obj.teacher else None

    if not student_name:
        return None
    if not group_name:
        return None

    bg_path = setting.background_image.path

    file_content = save_certificate_pdf(student_name, group_name, teacher_name, level, result.score, bg_path)
    if file_content is None:
        return None

    cert = Certificate(
        student_name=student_name,
        group_name=group_name,
        score=result.score,
        quiz_result=result,
    )
    cert.certificate_file.save(file_content.name, file_content, save=True)
    return cert


@login_required
def view_certificate(request, cert_id):
    cert = get_object_or_404(Certificate, id=cert_id)

    if not request.user.is_staff and not request.user.is_superuser:
        if cert.quiz_result and cert.quiz_result.student:
            if cert.quiz_result.student.user != request.user:
                messages.error(request, "Bu sertifikatni ko'rish huquqi yo'q!")
                return redirect('home')

    cert_setting = CertificateSetting.objects.filter(is_active=True).first()
    threshold = cert_setting.threshold_percentage if cert_setting else 50
    context = {
        'cert': cert,
        'threshold': threshold,
    }
    return render(request, 'groups/view_certificate.html', context)


@login_required
@user_passes_test(is_admin_user)
def generate_certificate_api(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Faqat POST so\'rov!'})

    try:
        data = json.loads(request.body)
        result_id = data.get('result_id')

        result = get_object_or_404(QuizResult, id=result_id)
        cert = generate_student_certificate(result)

        if cert:
            return JsonResponse({
                'success': True,
                'message': 'Sertifikat yaratildi!',
                'cert_id': cert.id,
                'cert_url': cert.certificate_file.url,
            })
        else:
            return JsonResponse({
                'success': False,
                'message': 'Sertifikat yaratilmadi. Sertifikat sozlamalarini tekshiring yoki ball yetarli emas.'
            })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
def my_certificates(request):
    student = Student.objects.filter(user=request.user).first()
    if not student:
        messages.error(request, "Siz student emassiz!")
        return redirect('home')

    certificates = Certificate.objects.filter(
        quiz_result__student=student
    ).order_by('-generated_at')

    if not certificates:
        certificates = Certificate.objects.filter(
            student_name__icontains=student.full_name
        ).order_by('-generated_at')

    context = {
        'certificates': certificates,
        'student': student,
    }
    return render(request, 'groups/my_certificates.html', context)


@login_required
def sertivkat_view(request):
    from django.templatetags.static import static
    setting = CertificateSetting.objects.filter(is_active=True).first()
    bg_image = None
    if setting and setting.background_image:
        try:
            if setting.background_image.storage.exists(setting.background_image.name):
                bg_image = setting.background_image.url
        except:
            bg_image = None
    context = {'bg_image': bg_image}
    return render(request, 'groups/sertivkat.html', context)


@login_required
@user_passes_test(is_admin_user)
def issue_certificates(request):
    setting = CertificateSetting.objects.filter(is_active=True).first()
    threshold = setting.threshold_percentage if setting else 50

    group_id = request.GET.get('group_id')
    specific_group = Group.objects.filter(id=group_id).first() if group_id else None

    existing_group_names = set(Group.objects.values_list('name', flat=True))
    all_results = QuizResult.objects.all().order_by('-submitted_at')
    if specific_group:
        all_results = all_results.filter(group_name_saved=specific_group.name)

    group_results_map = {}
    for r in all_results:
        gname = None
        if r.quiz_session and r.quiz_session.group:
            gname = r.quiz_session.group.name
        elif r.group_name_saved:
            gname = r.group_name_saved
        else:
            gname = "Noma'lum guruh"
        if gname not in group_results_map:
            group_results_map[gname] = {'group_name': gname, 'is_deleted': gname not in existing_group_names, 'results': []}
        group_results_map[gname]['results'].append(r)

    groups_data = []
    for gname, gdata in group_results_map.items():
        group = Group.objects.filter(name=gname).first()
        exam_config = GroupExamConfig.objects.filter(group=group).first() if group else None

        def get_grade(score_val):
            if exam_config and exam_config.grading_enabled:
                if score_val < exam_config.low_threshold:
                    return 'red', exam_config.label_low
                elif score_val >= exam_config.high_threshold:
                    return 'green', exam_config.label_high
                else:
                    return 'yellow', exam_config.label_medium
            else:
                if score_val >= 70:
                    return 'green', 'Yuqori'
                elif score_val >= 50:
                    return 'yellow', "O'rta"
                else:
                    return 'red', 'Past'

        students_data = {}
        for r in gdata['results']:
            student_key = r.student_id or r.student_name_saved
            student_name = r.student_name_saved or (r.student.full_name if r.student else 'Noma\'lum')
            if student_key not in students_data:
                students_data[student_key] = {
                    'student_name': student_name, 'student_id': r.student_id,
                    'attempts': [], 'best_score': 0, 'total_attempts': 0,
                    'has_certificate': False, 'eligible': False,
                }
            students_data[student_key]['attempts'].append(r)
            students_data[student_key]['total_attempts'] += 1
            score_val = float(r.score)
            if score_val > students_data[student_key]['best_score']:
                students_data[student_key]['best_score'] = score_val

        for sdata in students_data.values():
            grade_class, grade_label = get_grade(sdata['best_score'])
            sdata['grade_class'] = grade_class
            sdata['grade'] = grade_label
            sdata['eligible'] = sdata['best_score'] >= threshold
            existing = Certificate.objects.filter(student_name__iexact=sdata['student_name'], group_name__iexact=gname).first()
            sdata['has_certificate'] = existing is not None

        groups_data.append({
            'group_name': gname, 'is_deleted': gdata['is_deleted'],
            'students': list(students_data.values()),
            'total_students': len(students_data),
        })

    return render(request, 'groups/issue_certificates.html', {
        'groups_data': groups_data,
        'total_groups': len(groups_data),
        'threshold': threshold,
        'has_background': setting and setting.background_image,
    })


@login_required
@user_passes_test(is_admin_user)
@require_http_methods(["POST"])
def bulk_generate_certificates_api(request):
    try:
        data = json.loads(request.body)
        group_name = data.get('group_name')

        setting = CertificateSetting.objects.filter(is_active=True).first()
        if not setting or not setting.background_image:
            return JsonResponse({'success': False, 'message': 'Sertifikat sozlamalari to\'liq emas. Fon rasmini yuklang!'})

        from django.db.models import Max

        best_results = {}
        qs = QuizResult.objects.all()
        if group_name:
            qs = qs.filter(group_name_saved=group_name)

        for r in qs:
            sname = r.student_name_saved or (r.student.full_name if r.student else None)
            if not sname:
                continue
            if sname not in best_results or r.score > best_results[sname].score:
                best_results[sname] = r

        generated = []
        errors = []
        for sname, r in best_results.items():
            if r.score < setting.threshold_percentage:
                continue
            existing = Certificate.objects.filter(
                student_name__iexact=sname, group_name__iexact=group_name
            ).first() if group_name else Certificate.objects.filter(student_name__iexact=sname).first()
            if existing:
                continue

            cert = generate_student_certificate(r)
            if cert:
                generated.append({'student_name': sname, 'cert_id': cert.id})
            else:
                errors.append(sname)

        return JsonResponse({
            'success': True,
            'generated': generated,
            'errors': errors,
            'count': len(generated),
        })
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': str(e)})


@staff_member_required
def admin_question_edit(request, pk):
    """Savolni tahrirlash - barcha savol turlari uchun"""
    question = get_object_or_404(QuizQuestion, id=pk)
    
    if request.method == 'POST':
        question_type = question.question_type
        category_id = request.POST.get('category')
        
        try:
            category = Category.objects.get(id=category_id)
            question.category = category
            question.points = int(request.POST.get('points', question.points))
            
            # ============ 1. FILL BLANK (Word Bank) ============
            if question_type == 'fill_blank':
                question.question_text = request.POST.get('fb_question_text', '').strip()
                question.correct_answer = request.POST.get('fb_correct_answer', '').strip()
                
                if not question.question_text or not question.correct_answer:
                    messages.error(request, "Savol matni va to'g'ri javobni kiriting!")
                    return redirect('admin_question_edit', pk=question.id)
                
                question.save()
                messages.success(request, "✅ Bo'sh joy (Word Bank) savoli tahrirlandi!")
                return redirect('admin_question_list')
            
            # ============ 2. FILL BLANK NO WORD ============
            elif question_type == 'fill_blank_no_word':
                question.question_text = request.POST.get('fbnw_question_text', '').strip()
                question.correct_answer = request.POST.get('fbnw_correct_answer', '').strip()
                
                if not question.question_text or not question.correct_answer:
                    messages.error(request, "Savol matni va to'g'ri javobni kiriting!")
                    return redirect('admin_question_edit', pk=question.id)
                
                question.save()
                messages.success(request, "✅ Bo'sh joy (variantlarsiz) savoli tahrirlandi!")
                return redirect('admin_question_list')
            
            # ============ 3. SENTENCE ARRANGEMENT ============
            elif question_type == 'sentence_arrangement':
                question.scrambled_words = request.POST.get('sa_scrambled_words', '').strip()
                question.correct_sentence = request.POST.get('sa_correct_sentence', '').strip()
                
                if not question.scrambled_words or not question.correct_sentence:
                    messages.error(request, "Barcha maydonlarni to'ldiring!")
                    return redirect('admin_question_edit', pk=question.id)
                
                question.save()
                messages.success(request, "✅ So'z tartibi savoli tahrirlandi!")
                return redirect('admin_question_list')
            
            # ============ 4. READING COMPREHENSION ============
            elif question_type == 'reading_comprehension':
                if question.reading_text:
                    return redirect('reading_text_edit', pk=question.reading_text.id)
                messages.error(request, "Matn topilmadi!")
                return redirect('admin_question_edit', pk=question.id)
            
            # ============ 5. TRUE/FALSE ============
            elif question_type == 'true_false':
                question.question_text = request.POST.get('tf_question_text', '').strip()
                question.correct_answer = request.POST.get('tf_correct_answer', '').strip()
                
                if not question.question_text:
                    messages.error(request, "Savol matnini kiriting!")
                    return redirect('admin_question_edit', pk=question.id)
                if not question.correct_answer:
                    messages.error(request, "To'g'ri javobni tanlang!")
                    return redirect('admin_question_edit', pk=question.id)
                
                question.save()
                messages.success(request, "✅ To'g'ri/Noto'g'ri savoli tahrirlandi!")
                return redirect('admin_question_list')
            
            # ============ 6. MULTIPLE CHOICE ============
            elif question_type == 'multiple_choice':
                question.question_text = request.POST.get('mc_question_text', '').strip()
                options = []
                
                for key, value in request.POST.items():
                    if key.startswith('option_') and value.strip():
                        options.append(value.strip())
                
                correct_option_index = request.POST.get('correct_option')
                if correct_option_index and correct_option_index.isdigit():
                    correct_option_index = int(correct_option_index)
                else:
                    correct_option_index = None
                
                if not question.question_text:
                    messages.error(request, "Savol matnini kiriting!")
                elif len(options) < 2:
                    messages.error(request, "Kamida 2 ta variant kiriting!")
                elif correct_option_index is None or correct_option_index >= len(options):
                    messages.error(request, "To'g'ri variantni tanlang!")
                else:
                    question.correct_answer = options[correct_option_index]
                    question.scrambled_words = json.dumps(options)
                    question.save()
                    messages.success(request, "✅ Test varianti savoli tahrirlandi!")
                    return redirect('admin_question_list')
            
            # ============ 7. UNDERLINE CORRECT ============
            elif question_type == 'underline_correct':
                sentence_text = request.POST.get('uc_sentence_text', '').strip()
                correct_answer = request.POST.get('uc_correct_answer', '').strip()
                
                if not sentence_text:
                    messages.error(request, "Gap matnini kiriting!")
                elif not correct_answer:
                    messages.error(request, "To'g'ri javobni kiriting!")
                else:
                    options = []
                    if '/' in sentence_text:
                        parts = sentence_text.split('/')
                        for part in parts:
                            cleaned = part.strip()
                            if cleaned:
                                options.append(cleaned)
                    else:
                        options = [sentence_text]
                    
                    question.question_text = sentence_text
                    question.correct_answer = correct_answer
                    question.scrambled_words = json.dumps(options)
                    question.save()
                    messages.success(request, "✅ To'g'ri so'zni tanlash savoli tahrirlandi!")
                    return redirect('admin_question_list')
            
            # ============ 8. MATCHING ============
            elif question_type == 'matching':
                instruction = request.POST.get('matching_instruction', '').strip()
                left_items = []
                right_items = []
                matches = {}
                
                for key, value in request.POST.items():
                    if key.startswith('left_item_') and value.strip():
                        left_items.append(value.strip())
                
                for key, value in request.POST.items():
                    if key.startswith('right_item_') and value.strip():
                        right_items.append(value.strip())
                
                for key, value in request.POST.items():
                    if key.startswith('match_answer_') and value:
                        idx = int(key.split('_')[-1])
                        matches[str(idx + 1)] = value
                
                if len(left_items) < 2:
                    messages.error(request, "Kamida 2 ta chap tomon elementi kiriting!")
                elif len(right_items) < 2:
                    messages.error(request, "Kamida 2 ta o'ng tomon elementi kiriting!")
                elif len(matches) != len(left_items):
                    messages.error(request, f"Barcha {len(left_items)} ta elementga javob tanlang!")
                else:
                    matching_data = {'left': left_items, 'right': right_items}
                    question.question_text = instruction
                    question.correct_answer = json.dumps(matches)
                    question.scrambled_words = json.dumps(matching_data)
                    question.save()
                    messages.success(request, "✅ Moslashtirish savoli tahrirlandi!")
                    return redirect('admin_question_list')
            
            # ============ 9. CLOZE MULTIPLE BLANKS ============
            elif question_type == 'cloze_multiple_blanks':
                cloze_text = request.POST.get('cloze_question_text', '').strip()
                blank_options_json = request.POST.get('cloze_blank_options', '{}')
                correct_answers_json = request.POST.get('cloze_correct_answers', '{}')
                
                try:
                    blank_options = json.loads(blank_options_json)
                    correct_answers = json.loads(correct_answers_json)
                except json.JSONDecodeError:
                    blank_options = {}
                    correct_answers = {}
                
                if not cloze_text:
                    messages.error(request, "Matnni kiriting!")
                elif not blank_options or len(blank_options) == 0:
                    messages.error(request, "Hech qanday bo'sh joy aniqlanmadi!")
                else:
                    question.question_text = cloze_text
                    question.correct_answer = json.dumps(correct_answers)
                    question.blank_options = blank_options
                    question.scrambled_words = json.dumps(blank_options)
                    question.save()
                    messages.success(request, "✅ Ko'p bo'sh joy savoli tahrirlandi!")
                    return redirect('admin_question_list')
            
            # ============ 10. COMPLETE THE WORDS ============
            elif question_type == 'complete_the_words':
                sentence_text = request.POST.get('ctw_sentence_text', '').strip()
                correct_answer = request.POST.get('ctw_correct_answer', '').strip().lower()
                
                if not sentence_text:
                    messages.error(request, "Gap matnini kiriting!")
                elif not correct_answer:
                    messages.error(request, "To'g'ri javobni kiriting!")
                else:
                    question.question_text = sentence_text
                    question.correct_answer = correct_answer
                    question.save()
                    messages.success(request, "✅ So'zlarni to'ldirish savoli tahrirlandi!")
                    return redirect('admin_question_list')
            
            elif question_type == 'writing':
                topic = request.POST.get('w_topic', '').strip()
                if not topic:
                    messages.error(request, "Mavzu matnini kiriting!")
                else:
                    question.question_text = topic
                    question.correct_answer = ''
                    question.save()
                    messages.success(request, "✅ Yozma ish (Writing) savoli tahrirlandi!")
                    return redirect('admin_question_list')

            elif question_type == 'speaking':
                topic = request.POST.get('s_topic', '').strip()
                if not topic:
                    messages.error(request, "Mavzu matnini kiriting!")
                else:
                    question.question_text = topic
                    question.correct_answer = ''
                    question.save()
                    messages.success(request, "✅ Og'zaki (Speaking) savoli tahrirlandi!")
                    return redirect('admin_question_list')

            else:
                messages.error(request, f"Noto'g'ri savol turi: {question_type}")
                
        except Category.DoesNotExist:
            messages.error(request, "Kategoriya topilmadi!")
        except Exception as e:
            messages.error(request, f'Xatolik: {str(e)}')
        
        return redirect('admin_question_edit', pk=question.id)
    
    # ============ GET so'rovi - tahrirlash formasini ko'rsatish ============
    categories = Category.objects.all()
    
    # Savol turiga qarab qo'shimcha ma'lumotlarni tayyorlash
    if question.question_type == 'multiple_choice' and question.scrambled_words:
        try:
            question.options_list = json.loads(question.scrambled_words)
        except:
            question.options_list = []
    
    elif question.question_type == 'matching' and question.scrambled_words:
        try:
            matching_data = json.loads(question.scrambled_words)
            question.matching_left = matching_data.get('left', [])
            question.matching_right = matching_data.get('right', [])
        except:
            question.matching_left = []
            question.matching_right = []
        try:
            question.matching_matches = json.loads(question.correct_answer) if question.correct_answer else {}
        except:
            question.matching_matches = {}
    
    elif question.question_type == 'cloze_multiple_blanks':
        try:
            question.cloze_blanks = json.loads(question.blank_options) if question.blank_options else {}
            question.cloze_correct_answers = json.loads(question.correct_answer) if question.correct_answer else {}
        except:
            question.cloze_blanks = {}
            question.cloze_correct_answers = {}
    
    elif question.question_type == 'underline_correct' and question.scrambled_words:
        try:
            question.underline_options = json.loads(question.scrambled_words)
        except:
            question.underline_options = []
    
    elif question.question_type == 'fill_blank' and question.correct_answer:
        if '|' in question.correct_answer:
            question.fill_blank_variants = [v.strip() for v in question.correct_answer.split('|')]
        else:
            question.fill_blank_variants = [question.correct_answer]
    
    return render(request, 'groups/admin_question_edit.html', {
        'question': question,
        'categories': categories,
    })


# ============================================================
# SPEAKING (Og'zaki) BAHOLASH
# ============================================================

@login_required
def speaking_review(request, group_id):
    if not is_admin_user(request.user) and not is_teacher_user(request.user):
        messages.error(request, 'Sizda bu sahifani ko\'rish huquqi yo\'q!')
        return redirect('login')

    if is_teacher_user(request.user) and not is_admin_user(request.user):
        teacher = request.user.teacher_profile
        if not teacher.is_active:
            messages.error(request, 'Profilingiz faol emas!')
            return redirect('teacher_panel')
        group = get_object_or_404(Group, id=group_id)
        if not teacher.all_groups and not teacher.groups.filter(id=group.id).exists():
            messages.error(request, 'Siz bu guruhga biriktirilmagansiz!')
            return redirect('teacher_panel')
    else:
        group = get_object_or_404(Group, id=group_id)
    results = QuizResult.objects.filter(quiz_session__group=group).order_by('-submitted_at')
    questions = QuizQuestion.objects.filter(category__group_categories__group=group, question_type='speaking').distinct()

    speaking_entries = []
    seen_combinations = set()

    for result in results:
        if not result.answers:
            continue
        for qid_str, ans_data in result.answers.items():
            if isinstance(ans_data, dict) and ans_data.get('type') == 'speaking':
                try:
                    question = QuizQuestion.objects.get(id=int(qid_str))
                except (QuizQuestion.DoesNotExist, ValueError):
                    continue
                combo = (result.student.id, question.id)
                if combo in seen_combinations:
                    continue
                speaking_entries.append({
                    'result_id': result.id,
                    'question_id': question.id,
                    'student_id': result.student.id,
                    'student_name': result.student.full_name,
                    'topic': question.question_text,
                    'answer': ans_data.get('user_answer', ''),
                    'graded': ans_data.get('graded', False),
                    'earned_points': ans_data.get('earned_points', 0),
                    'max_points': question.points,
                    'submitted_at': result.submitted_at,
                })
                seen_combinations.add(combo)

    for student in group.students.all().select_related('user'):
        for q in questions:
            if (student.id, q.id) not in seen_combinations:
                speaking_entries.append({
                    'result_id': None,
                    'question_id': q.id,
                    'student_id': student.id,
                    'student_name': student.full_name,
                    'topic': q.question_text,
                    'answer': '',
                    'graded': False,
                    'earned_points': 0,
                    'max_points': q.points,
                    'submitted_at': None,
                })

    return render(request, 'groups/admin_speaking_review.html', {
        'group': group,
        'speaking_entries': speaking_entries,
        'questions': questions,
    })


@login_required
@csrf_exempt
def speaking_save_score_api(request):
    if not is_admin_user(request.user) and not is_teacher_user(request.user):
        return JsonResponse({'success': False, 'message': 'Huquq yo\'q'})

    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'POST kerak'})
    try:
        data = json.loads(request.body)
        student_id = data.get('student_id')
        question_id = data.get('question_id')
        group_id = data.get('group_id')
        earned_points = float(data.get('earned_points', 0))

        student = get_object_or_404(Student, id=student_id)
        question = get_object_or_404(QuizQuestion, id=question_id, question_type='speaking')
        group = get_object_or_404(Group, id=group_id)
        max_points = question.points
        earned_points = max(0, min(earned_points, max_points))

        quiz_session = QuizSession.objects.filter(group=group, is_active=True).first()
        if not quiz_session:
            quiz_session = QuizSession.objects.filter(group=group).order_by('-started_at').first()
        if not quiz_session:
            quiz_session = QuizSession.objects.create(
                group=group, is_active=False, started_at=timezone.now(), created_by=request.user
            )

        latest = QuizResult.objects.filter(
            student=student, quiz_session=quiz_session
        ).order_by('-attempt_number', '-id').first()

        if latest:
            result = latest
            created = False
        else:
            result, created = QuizResult.objects.get_or_create(
                student=student, quiz_session=quiz_session, attempt_number=1,
                defaults={'score': 0, 'total_questions': max_points, 'answers': {}}
            )

        if not result.answers:
            result.answers = {}

        qid_str = str(question.id)
        result.answers[qid_str] = {
            'type': 'speaking',
            'earned_points': earned_points,
            'max_points': max_points,
            'graded': True,
            'is_correct': earned_points > 0,
        }

        total_score = 0
        total_possible = 0
        qids_in_answers = [int(k) for k in result.answers.keys() if str(k).isdigit()]
        questions = QuizQuestion.objects.filter(id__in=qids_in_answers) if qids_in_answers else QuizQuestion.objects.none()
        for q in questions:
            sqid = str(q.id)
            total_possible += q.points
            a = result.answers[sqid]
            if isinstance(a, dict):
                if a.get('type') in ('writing', 'speaking'):
                    total_score += a.get('earned_points', 0)
                elif 'blanks' in a:
                    pts_per = round(q.points / max(a.get('blanks_total', 1), 1), 2)
                    total_score += a.get('blanks_correct', 0) * pts_per
                elif a.get('is_correct'):
                    total_score += q.points

        result.score = round((total_score / total_possible) * 100, 1) if total_possible > 0 else 0
        result.total_questions = total_possible
        result.save()

        AssessmentScore.objects.update_or_create(
            student=student,
            group=group,
            assessment_type='speaking',
            defaults={
                'score': int(earned_points),
                'added_by': request.user,
                'student_name_saved': student.full_name,
                'group_name_saved': group.name,
            }
        )

        return JsonResponse({
            'success': True,
            'earned': earned_points,
            'max_points': max_points,
            'score': result.score,
        })

    except Exception as e:
        import traceback
        traceback.print_exc()
        return JsonResponse({'success': False, 'message': str(e)})


# ============================================================
# NATIJALARNI EXPORT QILISH (CSV/Excel)
# ============================================================

@login_required
@user_passes_test(is_admin_user)
def export_results_csv(request, group_id):
    import csv
    from django.http import HttpResponse

    group = get_object_or_404(Group, id=group_id)

    latest_results = QuizResult.objects.filter(
        quiz_session__group=group
    ).values('student_id').annotate(
        latest_id=Max('id')
    ).values_list('latest_id', flat=True)

    all_results = QuizResult.objects.filter(
        id__in=latest_results
    ).select_related('student__user').order_by('-submitted_at')

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{group.name}_natijalar.csv"'
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow(['#', 'Student', 'Username', 'Ball', 'Maks. Ball', 'Foiz', 'Sana'])

    for i, result in enumerate(all_results, 1):
        total_possible = result.total_questions or 0
        score_pct = float(result.score)
        raw_score = round((score_pct / 100) * total_possible, 1) if total_possible > 0 else 0
        writer.writerow([
            i,
            result.student.full_name,
            result.student.user.username,
            raw_score,
            total_possible,
            f'{score_pct}%',
            result.submitted_at.strftime('%Y-%m-%d %H:%M:%S'),
        ])

    return response


@login_required
@user_passes_test(is_admin_user)
def export_results_excel(request, group_id):
    group = get_object_or_404(Group, id=group_id)
    from django.http import HttpResponse

    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        messages.error(request, "openpyxl kutubxonasi o'rnatilmagan. pip install openpyxl")
        return redirect('quiz_results', group_id=group_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Natijalar"

    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")

    headers = ['#', 'Student', 'Username', 'Ball', 'Maks. Ball', 'Foiz', 'Sana']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    latest_results = QuizResult.objects.filter(
        quiz_session__group=group
    ).values('student_id').annotate(
        latest_id=Max('id')
    ).values_list('latest_id', flat=True)

    all_results = QuizResult.objects.filter(
        id__in=latest_results
    ).select_related('student__user').order_by('-submitted_at')

    for i, result in enumerate(all_results, 1):
        total_possible = result.total_questions or 0
        score_pct = float(result.score)
        raw_score = round((score_pct / 100) * total_possible, 1) if total_possible > 0 else 0
        ws.cell(row=i+1, column=1, value=i)
        ws.cell(row=i+1, column=2, value=result.student.full_name)
        ws.cell(row=i+1, column=3, value=result.student.user.username)
        ws.cell(row=i+1, column=4, value=raw_score)
        ws.cell(row=i+1, column=5, value=total_possible)
        ws.cell(row=i+1, column=6, value=f'{score_pct}%')
        ws.cell(row=i+1, column=7, value=result.submitted_at.strftime('%Y-%m-%d %H:%M:%S'))

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 22

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{group.name}_natijalar.xlsx"'
    wb.save(response)
    return response


# ============================================================
# STATISTIKA DIAGRAMMALAR BILAN
# ============================================================

@login_required
@user_passes_test(is_admin_user)
def quiz_statistics(request, group_id):
    group = get_object_or_404(Group, id=group_id)

    latest_results = QuizResult.objects.filter(
        quiz_session__group=group
    ).values('student_id').annotate(
        latest_id=Max('id')
    ).values_list('latest_id', flat=True)

    results = QuizResult.objects.filter(
        id__in=latest_results
    ).select_related('student__user').order_by('-submitted_at')

    score_ranges = {'0-20': 0, '21-40': 0, '41-60': 0, '61-80': 0, '81-100': 0}
    student_scores = []
    labels = []
    score_values = []
    colors = []

    for r in results:
        score = float(r.score)
        name = r.student.full_name
        student_scores.append({'name': name, 'score': score})
        labels.append(name)
        score_values.append(score)

        if score <= 20: score_ranges['0-20'] += 1
        elif score <= 40: score_ranges['21-40'] += 1
        elif score <= 60: score_ranges['41-60'] += 1
        elif score <= 80: score_ranges['61-80'] += 1
        else: score_ranges['81-100'] += 1

    # Baholash tizimi
    exam_config = GroupExamConfig.objects.filter(group=group).first()
    for s in student_scores:
        score = s['score']
        if exam_config and exam_config.grading_enabled:
            if score < exam_config.low_threshold:
                s['grade_class'] = 'red'
                s['grade'] = exam_config.label_low
            elif score >= exam_config.high_threshold:
                s['grade_class'] = 'green'
                s['grade'] = exam_config.label_high
            else:
                s['grade_class'] = 'yellow'
                s['grade'] = exam_config.label_medium
        else:
            if score >= 70:
                s['grade_class'] = 'green'
                s['grade'] = 'Yuqori'
            elif score >= 50:
                s['grade_class'] = 'yellow'
                s['grade'] = "O'rta"
            else:
                s['grade_class'] = 'red'
                s['grade'] = 'Past'

    total = len(results)
    avg_score = round(sum(score_values) / total, 1) if total > 0 else 0
    max_score = max(score_values) if score_values else 0
    min_score = min(score_values) if score_values else 0
    above_70 = sum(1 for s in score_values if s >= 70)
    below_40 = sum(1 for s in score_values if s < 40)

    return render(request, 'groups/quiz_statistics.html', {
        'group': group,
        'total_students': total,
        'avg_score': avg_score,
        'max_score': max_score,
        'min_score': min_score,
        'above_70': above_70,
        'below_40': below_40,
        'score_ranges': json.dumps(score_ranges),
        'labels': json.dumps(labels),
        'score_values': json.dumps(score_values),
        'student_scores': student_scores,
        'exam_config': exam_config,
    })


# ============================================================
# QURILMA NAZORATI
# ============================================================


@login_required
@user_passes_test(is_admin_user)
def device_monitor(request):
    """Qurilma nazorati - admin barcha qurilmalarni ko'radi"""
    from datetime import timedelta
    now = timezone.now()

    devices = Device.objects.select_related('student__user', 'group').order_by('-last_seen')

    # Online deb hisoblash: oxirgi 2 daqiqa ichida heartbeat kelgan bo'lsa
    threshold = now - timedelta(minutes=2)
    online_count = devices.filter(last_seen__gte=threshold).count()

    # AJAX so'rov bo'lsa JSON qaytarish
    if request.GET.get('ajax') == '1':
        dev_list = []
        for d in devices:
            is_online = d.last_seen and d.last_seen >= threshold
            dev_list.append({
                'device_id': d.device_id,
                'name': d.name or '',
                'student_name': d.student.full_name if d.student else None,
                'group_name': d.group.name if d.group else None,
                'platform': d.platform or '',
                'ip_address': d.ip_address or '',
                'last_seen': d.last_seen.isoformat() if d.last_seen else None,
                'is_active': is_online,
            })
        return JsonResponse({'success': True, 'devices': dev_list, 'online_count': online_count, 'total_count': devices.count()})

    # Guruh bo'yicha guruhlash
    groups_map = {}
    for d in devices:
        gname = d.group.name if d.group else "Guruhsiz"
        if gname not in groups_map:
            groups_map[gname] = {'group_name': gname, 'devices': []}
        groups_map[gname]['devices'].append(d)

    total_devices = devices.count()

    return render(request, 'groups/device_monitor.html', {
        'devices': devices,
        'groups_map': dict(groups_map),
        'total_active': online_count,
        'online_count': online_count,
        'total_devices': total_devices,
        'threshold': threshold,  # MUHIM: threshold ni templatega yuboramiz
    })

@csrf_exempt
def device_register_api(request):
    """Qurilma ro'yxatdan o'tadi va heart beat yuboradi"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Faqat POST'}, status=405)

    try:
        data = json.loads(request.body)
        device_id = data.get('device_id')
        if not device_id:
            return JsonResponse({'success': False, 'message': 'device_id kerak'})

        # Studentni topish (agar login bo'lsa)
        student = None
        group = None
        if request.user.is_authenticated:
            try:
                student = request.user.student_profile
                group = student.group
            except:
                pass

        ip = request.META.get('REMOTE_ADDR', '')
        # Proxy orqali bo'lsa
        xff = request.META.get('HTTP_X_FORWARDED_FOR')
        if xff:
            ip = xff.split(',')[0].strip()

        device, created = Device.objects.update_or_create(
            device_id=device_id,
            defaults={
                'student': student,
                'group': group,
                'user_agent': data.get('user_agent', ''),
                'ip_address': ip,
                'platform': data.get('platform', ''),
                'screen_resolution': data.get('screen_resolution', ''),
                'is_active': True,
            }
        )
        return JsonResponse({'success': True, 'created': created, 'device_id': device.device_id})

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@csrf_exempt
def device_offline_api(request):
    """Qurilmani offline qilish (tab yopilganda)"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Faqat POST'}, status=405)
    try:
        data = json.loads(request.body)
        device_id = data.get('device_id')
        if not device_id:
            return JsonResponse({'success': False, 'message': 'device_id kerak'})
        from datetime import timedelta
        Device.objects.filter(device_id=device_id).update(is_active=False, last_seen=timezone.now() - timedelta(minutes=10))
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@user_passes_test(is_admin_user)
def device_rename_api(request):
    """Admin qurilma nomini o'zgartiradi"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Faqat POST'}, status=405)
    try:
        data = json.loads(request.body)
        device = get_object_or_404(Device, device_id=data.get('device_id'))
        device.name = data.get('name', '')
        device.save(update_fields=['name'])
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@user_passes_test(is_admin_user)
def device_delete_api(request):
    """Admin qurilmani o'chiradi"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Faqat POST'}, status=405)
    try:
        data = json.loads(request.body)
        device = get_object_or_404(Device, device_id=data.get('device_id'))
        device.delete()
        return JsonResponse({'success': True})
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})


@login_required
@user_passes_test(is_admin_user)
def device_history(request):
    """Qurilma tarixi - qaysi student qaysi qurilmada ishlagan"""
    devices = Device.objects.select_related('student__user', 'group').order_by('-last_seen')
    students_map = {}
    for d in devices:
        if d.student:
            key = d.student.id
            if key not in students_map:
                students_map[key] = {
                    'student': d.student,
                    'devices': []
                }
            students_map[key]['devices'].append(d)

    total_students = len(students_map)
    total_devices = devices.count()

    return render(request, 'groups/device_history.html', {
        'students_map': dict(students_map),
        'total_students': total_students,
        'total_devices': total_devices,
    })


# ============================================================
# ARXIV
# ============================================================


@login_required
@user_passes_test(is_admin_user)
def results_archive(request):
    """Arxiv - barcha studentlarning barcha natijalari"""

    # Mavjud guruhlar
    existing_group_names = set(Group.objects.values_list('name', flat=True))

    # Barcha natijalarni guruh nomi bo'yicha guruhlash
    group_results_map = {}
    all_results = QuizResult.objects.all().order_by('-submitted_at')

    for r in all_results:
        gname = None
        if r.quiz_session and r.quiz_session.group:
            gname = r.quiz_session.group.name
        elif r.group_name_saved:
            gname = r.group_name_saved
        elif r.quiz_session:
            gname = f"Sessiya #{r.quiz_session.id}"
        else:
            gname = "Noma'lum guruh"

        if gname not in group_results_map:
            group_results_map[gname] = {
                'group_name': gname,
                'is_deleted': gname not in existing_group_names,
                'results': [],
            }
        group_results_map[gname]['results'].append(r)

    # Har bir guruh uchun studentlarni guruhlash
    groups_data = []
    for gname, gdata in group_results_map.items():
        # Guruh sozlamalarini olish (agar guruh mavjud bo'lsa)
        group = Group.objects.filter(name=gname).first()
        exam_config = GroupExamConfig.objects.filter(group=group).first() if group else None

        def get_grade(score_val):
            if exam_config and exam_config.grading_enabled:
                if score_val < exam_config.low_threshold:
                    return 'red', exam_config.label_low
                elif score_val >= exam_config.high_threshold:
                    return 'green', exam_config.label_high
                else:
                    return 'yellow', exam_config.label_medium
            else:
                if score_val >= 70:
                    return 'green', 'Yuqori'
                elif score_val >= 50:
                    return 'yellow', "O'rta"
                else:
                    return 'red', 'Past'

        students_data = {}
        for r in gdata['results']:
            student_key = r.student_id or r.student_name_saved
            student_name = r.student_name_saved or (r.student.full_name if r.student else 'Noma\'lum')
            if student_key not in students_data:
                device_info = None
                if r.student_id:
                    device = Device.objects.filter(student_id=r.student_id).order_by('-last_seen').first()
                    if device:
                        device_info = device.name if device.name else device.device_id[:15] + '...'
                students_data[student_key] = {
                    'student_name': student_name,
                    'student_id': r.student_id,
                    'device_name': device_info,
                    'attempts': [],
                    'best_score': 0,
                    'worst_score': 100,
                    'total_attempts': 0,
                }
            students_data[student_key]['attempts'].append(r)
            students_data[student_key]['total_attempts'] += 1
            score_val = float(r.score)
            if score_val > students_data[student_key]['best_score']:
                students_data[student_key]['best_score'] = score_val
            if score_val < students_data[student_key]['worst_score']:
                students_data[student_key]['worst_score'] = score_val

        for sdata in students_data.values():
            grade_class, grade_label = get_grade(sdata['best_score'])
            sdata['grade_class'] = grade_class
            sdata['grade'] = grade_label

        groups_data.append({
            'group_name': gname,
            'is_deleted': gdata['is_deleted'],
            'students': list(students_data.values()),
            'total_students': len(students_data),
        })

    return render(request, 'groups/results_archive.html', {
        'groups_data': groups_data,
        'total_groups': len(groups_data),
        'total_results': all_results.count(),
    })


# ============================================================
# O'QITUVCHI (TEACHER)
# ============================================================


def is_teacher_user(user):
    return user.is_authenticated and hasattr(user, 'teacher_profile')


@csrf_exempt
def teacher_login(request):
    from django.contrib.auth.forms import AuthenticationForm
    if request.user.is_authenticated:
        if is_teacher_user(request.user):
            return redirect('teacher_panel')
        if is_admin_user(request.user):
            return redirect('admin_panel')
        return redirect('student_panel')

    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=username, password=password)
            if user is not None and is_teacher_user(user):
                login(request, user)
                messages.success(request, f'Xush kelibsiz, {user.get_full_name() or user.username}!')
                return redirect('teacher_panel')
            elif user is not None:
                messages.error(request, 'Siz o\'qituvchi sifatida ro\'yxatdan o\'tmagansiz!')
            else:
                messages.error(request, 'Username yoki parol xato!')
        else:
            messages.error(request, 'Username yoki parol xato!')
    else:
        form = AuthenticationForm()

    return render(request, 'groups/teacher_login.html', {'form': form})


@login_required
def teacher_panel(request):
    if not is_teacher_user(request.user):
        messages.error(request, 'Sizda bu sahifani ko\'rish huquqi yo\'q!')
        return redirect('login')

    teacher = request.user.teacher_profile
    if teacher.all_groups:
        groups = Group.objects.all().order_by('name')
    else:
        groups = teacher.groups.all().order_by('name')

    return render(request, 'groups/teacher_panel.html', {
        'teacher': teacher,
        'groups': groups,
    })


@login_required
def teacher_group_view(request, group_id):
    if not is_teacher_user(request.user):
        messages.error(request, 'Sizda bu sahifani ko\'rish huquqi yo\'q!')
        return redirect('login')

    teacher = request.user.teacher_profile
    group = get_object_or_404(Group, id=group_id)

    # Faqat o'z guruhlariga kirish
    if not teacher.all_groups and not teacher.groups.filter(id=group.id).exists():
        messages.error(request, 'Siz bu guruhga biriktirilmagansiz!')
        return redirect('teacher_panel')

    students = Student.objects.filter(group=group).order_by('user__first_name')

    return render(request, 'groups/teacher_group.html', {
        'teacher': teacher,
        'group': group,
        'students': students,
    })

@login_required
@user_passes_test(is_admin_user)
def teacher_score_logs(request):
    logs = []
    
    # TeacherScoreLog larni qo'shish
    teacher_logs = TeacherScoreLog.objects.select_related('teacher__user', 'student__user').all()
    for log in teacher_logs:
        logs.append({
            'type': 'teacher_log',
            'teacher_name': log.teacher.user.get_full_name() or log.teacher.user.username,
            'student_name': log.student_name_saved or (log.student.full_name if log.student else 'Noma\'lum'),
            'group_name': log.group_name_saved or (log.student.group.name if log.student and log.student.group else '-'),
            'score': log.score_added,
            'comment': log.comment or '',
            'created_at': log.created_at,
        })
    
    # QuizResult larni qo'shish (test natijalari)
    quiz_results = QuizResult.objects.select_related('student__user', 'quiz_session').all()
    for qr in quiz_results:
        logs.append({
            'type': 'quiz',
            'teacher_name': 'Avtomatik',
            'student_name': qr.student_name_saved or (qr.student.full_name if qr.student else 'Noma\'lum'),
            'group_name': qr.group_name_saved or (qr.student.group.name if qr.student and qr.student.group else '-'),
            'score': qr.score,
            'comment': f"Test: {qr.total_questions} ta savol",
            'created_at': qr.submitted_at,
        })

    # AssessmentScore larni qo'shish (Speaking va Writing)
    assessment_logs = AssessmentScore.objects.select_related('added_by', 'student__user').all()
    for a in assessment_logs:
        added_by_name = a.added_by.get_full_name() or a.added_by.username if a.added_by else 'Noma\'lum'
        assessment_type_display = dict(AssessmentScore.ASSESSMENT_TYPES).get(a.assessment_type, a.assessment_type)
        
        logs.append({
            'type': 'assessment',
            'assessment_type': a.assessment_type,
            'assessment_type_display': assessment_type_display,
            'teacher_name': added_by_name,
            'student_name': a.student_name_saved or (a.student.full_name if a.student else 'Noma\'lum'),
            'group_name': a.group_name_saved or (a.student.group.name if a.student and a.student.group else '-'),
            'score': a.score,
            'comment': assessment_type_display,
            'created_at': a.created_at,
        })
    
    # Vaqt bo'yicha saralash (eng yangisi birinchi)
    logs.sort(key=lambda x: x['created_at'], reverse=True)
    
    return render(request, 'groups/teacher_score_logs.html', {'logs': logs})
@login_required
@user_passes_test(is_admin_user)
def teacher_list(request):
    teachers = Teacher.objects.select_related('user').all().order_by('-created_at')
    return render(request, 'groups/teacher_list.html', {'teachers': teachers})


@login_required
@user_passes_test(is_admin_user)
def teacher_add(request):
    groups = Group.objects.all().order_by('name')
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        group_ids = request.POST.getlist('groups')
        all_groups = request.POST.get('all_groups') == 'on'
        is_active = request.POST.get('is_active') == 'on'

        if not username or not password:
            messages.error(request, 'Username va parol majburiy!')
        elif User.objects.filter(username=username).exists():
            messages.error(request, 'Bu username band!')
        else:
            try:
                with transaction.atomic():
                    user = User.objects.create_user(
                        username=username,
                        password=password,
                        first_name=first_name,
                        last_name=last_name,
                    )
                    teacher = Teacher.objects.create(user=user, all_groups=all_groups, is_active=is_active)
                    if not all_groups and group_ids:
                        teacher.groups.set(Group.objects.filter(id__in=group_ids))
                    messages.success(request, f'O\'qituvchi "{user.get_full_name() or user.username}" yaratildi!')
                    return redirect('teacher_list')
            except Exception as e:
                messages.error(request, f'Xatolik: {str(e)}')

    return render(request, 'groups/teacher_form.html', {
        'groups': groups,
        'mode': 'add',
    })


@login_required
@user_passes_test(is_admin_user)
def teacher_edit(request, teacher_id):
    teacher = get_object_or_404(Teacher.objects.select_related('user'), id=teacher_id)
    groups = Group.objects.all().order_by('name')

    if request.method == 'POST':
        teacher.user.first_name = request.POST.get('first_name', '').strip()
        teacher.user.last_name = request.POST.get('last_name', '').strip()
        password = request.POST.get('password', '').strip()
        group_ids = request.POST.getlist('groups')
        all_groups = request.POST.get('all_groups') == 'on'
        is_active = request.POST.get('is_active') == 'on'

        teacher.all_groups = all_groups
        teacher.is_active = is_active
        teacher.save()

        teacher.user.save()
        if not all_groups and group_ids:
            teacher.groups.set(Group.objects.filter(id__in=group_ids))
        else:
            teacher.groups.clear()

        if password:
            teacher.user.set_password(password)
            teacher.user.save()

        messages.success(request, 'O\'qituvchi ma\'lumotlari yangilandi!')
        return redirect('teacher_list')

    return render(request, 'groups/teacher_form.html', {
        'teacher': teacher,
        'groups': groups,
        'mode': 'edit',
    })


@login_required
@user_passes_test(is_admin_user)
def teacher_delete(request, teacher_id):
    teacher = get_object_or_404(Teacher, id=teacher_id)
    name = teacher.user.get_full_name() or teacher.user.username
    teacher.user.delete()  # User va Teacher cascade bo'yicha o'chadi
    messages.success(request, f'O\'qituvchi "{name}" o\'chirildi!')
    return redirect('teacher_list')


# ============ BAHOLASH (Speaking & Written) ============

@login_required
@require_http_methods(["POST"])
def admin_save_assessment_api(request):
    if not is_admin_user(request.user):
        return JsonResponse({'success': False, 'message': 'Huquq yo\'q'})

    try:
        data = json.loads(request.body)
        student_id = data.get('student_id')
        assessment_type = data.get('assessment_type')
        score = int(data.get('score', 0))

        if assessment_type not in ['speaking', 'written']:
            return JsonResponse({'success': False, 'message': 'Noto\'g\'ri baholash turi'})

        if score < 1 or score > 100:
            return JsonResponse({'success': False, 'message': 'Ball 1-100 oralig\'ida bo\'lishi kerak'})

        student = get_object_or_404(Student, id=student_id)
        group = student.group

        assessment, created = AssessmentScore.objects.update_or_create(
            student=student,
            group=group,
            assessment_type=assessment_type,
            defaults={
                'score': score,
                'added_by': request.user,
                'student_name_saved': student.full_name,
                'group_name_saved': group.name,
            }
        )

        action = "qo'shildi" if created else "yangilandi"
        assessment_name = dict(AssessmentScore.ASSESSMENT_TYPES)[assessment_type]

        return JsonResponse({
            'success': True,
            'message': f'{student.full_name} uchun {assessment_name} bahosi ({score}) {action}',
            'created': created,
            'score': assessment.score,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})